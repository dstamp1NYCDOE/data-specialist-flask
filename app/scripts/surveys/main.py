"""
Main analysis function for student belongingness survey
UPDATED VERSION with set cover optimization integrated
"""
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from collections import defaultdict

from .utils import (
    identify_missing_students,
    calculate_weighted_scores,
    assign_risk_profiles,
    assign_tiers,
    rank_students_within_tiers,
    create_diverging_bar_data,
    preprocess_survey_responses
)


def analyze_survey(
    df,
    config,
    rosters_df=None,
    master_schedule_df=None,
    biographical_columns=None,
    handle_missing='flag',
    question_text_map=None
):
    """
    Main function to analyze survey data and generate Excel report.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Survey data with StudentID, FirstName, LastName, Counselor, year_in_hs, 
        FormID, and Q1-Q16 columns
    config : object
        Survey configuration object (e.g., from belongingness_config.py)
    rosters_df : pd.DataFrame, optional
        Student course rosters with columns: StudentID, Course, Section
    master_schedule_df : pd.DataFrame, optional
        Master schedule with columns: Course, Section, Room, Teacher1, Teacher2, Period
    biographical_columns : list, optional
        Additional biographical columns beyond Counselor and year_in_hs
    handle_missing : str, default 'flag'
        How to handle missing survey responses: 'flag' or 'impute'
    question_text_map : dict, optional
        Mapping of question numbers to full question text for charts
        
    Returns:
    --------
    BytesIO
        Excel file in memory ready to send via Flask response
    """
    
    # Initialize biographical columns
    if biographical_columns is None:
        biographical_columns = []
    
    # Core columns that should always be present
    core_bio_columns = ['Counselor', 'year_in_hs']
    all_bio_columns = core_bio_columns + biographical_columns
    
    # Validate required columns
    required_cols = ['StudentID', 'FirstName', 'LastName'] + core_bio_columns + ['FormID']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns: {missing_cols}")
    
    # Identify question columns based on config
    question_cols = config.get_question_columns()
    
    # Preprocess survey responses to handle multiple bubbles
    df = preprocess_survey_responses(df, question_cols)
    
    # Identify missing students (all question responses are NaN)
    missing_students_df = identify_missing_students(df, question_cols)
    
    # Create working dataframe (exclude missing students if flagged)
    if handle_missing == 'flag':
        working_df = df[~df['StudentID'].isin(missing_students_df['StudentID'])].copy()
        # For missing students, assign conservative score of 2 for section analysis
        missing_students_df_with_scores = missing_students_df.copy()
        for col in question_cols:
            missing_students_df_with_scores[col] = 2.0
        # Calculate scores for missing students too
        missing_students_df_with_scores = calculate_weighted_scores(missing_students_df_with_scores, config)
        missing_students_df_with_scores = calculate_dimension_subscores(missing_students_df_with_scores, config)
    else:
        working_df = df.copy()
        missing_students_df_with_scores = pd.DataFrame()
    
    # Handle missing values for students with partial responses
    if handle_missing == 'impute':
        for col in question_cols:
            if col in working_df.columns:
                mean_val = working_df[col].mean()
                working_df[col].fillna(mean_val, inplace=True)
    
    # Calculate weighted risk scores
    working_df = calculate_weighted_scores(working_df, config)
    
    # Assign risk profiles
    working_df = assign_risk_profiles(working_df, config)
    
    # Assign MTSS tiers (by counselor)
    working_df = assign_tiers(working_df, config)
    
    # Rank students within tiers
    working_df = rank_students_within_tiers(working_df, config)
    
    # Create dimension subscores for analysis
    working_df = calculate_dimension_subscores(working_df, config)
    
    # Generate Excel workbook
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Raw analyzed data
        working_df.to_excel(writer, sheet_name='Analysis', index=False)
        
        # Sheet 2: Missing students
        if not missing_students_df.empty:
            missing_students_df.to_excel(writer, sheet_name='Missing Responses', index=False)
        
        # Sheet 3-5: Tier lists
        create_tier_sheets(writer, working_df, config)
        
        # Sheet 6+: Visualization data (by year_in_hs and counselor only)
        viz_sheets = create_visualization_sheets(
            writer, 
            working_df, 
            all_bio_columns, 
            config,
            question_text_map
        )
        
        # Sheet: Section Analysis by Dimension
        if rosters_df is not None and master_schedule_df is not None:
            create_section_analysis_sheet(
                writer, 
                working_df, 
                missing_students_df_with_scores,
                rosters_df, 
                master_schedule_df, 
                config
            )
    
    # Reload workbook to add formatting
    output.seek(0)
    wb = load_workbook(output)
    
    # Format all sheets
    format_workbook(wb)
    
    # Save back to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def calculate_dimension_subscores(df, config):
    """Calculate subscores for different dimensions of belongingness."""
    if hasattr(config, 'dimension_groups'):
        for dimension_name, questions in config.dimension_groups.items():
            cols = [f'Q{q}' for q in questions if f'Q{q}' in df.columns]
            if cols:
                df[f'{dimension_name}_subscore'] = df[cols].sum(axis=1)
                df[f'{dimension_name}_avg'] = df[cols].mean(axis=1)
    return df


def create_tier_sheets(writer, df, config):
    """Create separate sheets for each tier with full student details."""
    tiers = ['Tier 3', 'Tier 2', 'Tier 1']
    
    for tier in tiers:
        tier_df = df[df['Tier'] == tier].copy()
        
        if not tier_df.empty:
            # Sort by rank
            tier_df = tier_df.sort_values('Tier_Rank')
            
            # Select columns to display
            display_cols = [
                'Tier_Rank', 'StudentID', 'FirstName', 'LastName', 
                'Counselor', 'year_in_hs', 'FormID',
                'Weighted_Risk_Score', 'Total_Score'
            ]
            
            # Add risk profile columns
            profile_cols = [col for col in tier_df.columns if col.startswith('Profile_')]
            display_cols.extend(profile_cols)
            
            # Add question scores
            question_cols = [col for col in tier_df.columns if col.startswith('Q') and col[1:].isdigit()]
            display_cols.extend(sorted(question_cols, key=lambda x: int(x[1:])))
            
            # Add dimension subscores if they exist
            subscore_cols = [col for col in tier_df.columns if '_subscore' in col or '_avg' in col]
            display_cols.extend(subscore_cols)
            
            # Filter to existing columns
            display_cols = [col for col in display_cols if col in tier_df.columns]
            
            # Add suggested interventions column
            tier_df['Suggested_Interventions'] = tier_df.apply(
                lambda row: config.get_intervention_suggestions(row), 
                axis=1
            )
            display_cols.append('Suggested_Interventions')
            
            tier_df[display_cols].to_excel(writer, sheet_name=tier, index=False)


def create_visualization_sheets(writer, df, bio_columns, config, question_text_map):
    """Create sheets with diverging bar chart data for year_in_hs and counselor only."""
    question_cols = config.get_question_columns()
    viz_sheets = []
    
    # Overall diverging bar data for each question
    overall_data = create_diverging_bar_data(df, question_cols, question_text_map)
    overall_data.to_excel(writer, sheet_name='Overall Distribution', index=False)
    viz_sheets.append(('Overall Distribution', 'overall'))
    
    # By year_in_hs and Counselor only
    for bio_col in ['year_in_hs', 'Counselor']:
        if bio_col in df.columns:
            bio_data = []
            for group_value in df[bio_col].dropna().unique():
                group_df = df[df[bio_col] == group_value]
                group_bar_data = create_diverging_bar_data(
                    group_df, 
                    question_cols, 
                    question_text_map
                )
                group_bar_data['Group'] = group_value
                bio_data.append(group_bar_data)
            
            if bio_data:
                combined_bio_data = pd.concat(bio_data, ignore_index=True)
                sheet_name = f'By {bio_col}'[:31]  # Excel sheet name limit
                combined_bio_data.to_excel(writer, sheet_name=sheet_name, index=False)
                viz_sheets.append((sheet_name, bio_col))
    
    return viz_sheets


def create_section_analysis_sheet(writer, survey_df, missing_students_df, rosters_df, master_schedule_df, config):
    """
    Create sheet analyzing class sections by dimension for Tier 2 intervention targeting.
    Now includes set cover optimization for intervention recommendations.
    
    Parameters:
    -----------
    writer : pd.ExcelWriter
        Excel writer object
    survey_df : pd.DataFrame
        Students who completed survey with dimension scores
    missing_students_df : pd.DataFrame
        Students who didn't complete survey (with conservative score of 2)
    rosters_df : pd.DataFrame
        Course rosters with StudentID, Course, Section
    master_schedule_df : pd.DataFrame
        Master schedule with Course, Section, Room, Teacher1, Teacher2, Period
    config : object
        Survey configuration
    """
    
    # Combine survey completers and non-completers
    # Ensure both have the same columns for concatenation
    if not missing_students_df.empty:
        # Get common columns
        common_cols = list(set(survey_df.columns) & set(missing_students_df.columns))
        combined_df = pd.concat([survey_df[common_cols], missing_students_df[common_cols]], ignore_index=True)
    else:
        combined_df = survey_df.copy()
    
    # Merge with rosters to get course/section info
    analysis_df = combined_df.merge(
        rosters_df[['StudentID', 'Course', 'Section']].drop_duplicates(),
        on='StudentID',
        how='inner'
    )
    
    # Merge with master schedule to get teacher/room/period
    master_schedule_clean = master_schedule_df.copy()
    # Convert Period to string to handle mixed types
    if 'Period' in master_schedule_clean.columns:
        master_schedule_clean['Period'] = master_schedule_clean['Period'].astype(str)
    
    analysis_df = analysis_df.merge(
        master_schedule_clean[['Course', 'Section', 'Room', 'Teacher1', 'Teacher2', 'Period']],
        on=['Course', 'Section'],
        how='left'
    ).fillna('')
    
    # Create section identifier
    analysis_df['Section_ID'] = analysis_df['Course'] + ' - ' + analysis_df['Section']
    
    results = []
    
    # Store dimension priority scores for composite calculation
    dimension_priorities = {}
    
    # Analyze overall belongingness (using Weighted_Risk_Score)
    if 'Weighted_Risk_Score' in analysis_df.columns:
        section_stats = calculate_section_stats(
            analysis_df,
            'Section_ID',
            'Weighted_Risk_Score',
            invert_score=True,  # Lower weighted risk score = better
            threshold=None  # We'll calculate % below a certain percentile instead
        )
        
        for _, row in section_stats.iterrows():
            results.append({
                'Dimension': 'Overall Belongingness',
                'Course': analysis_df[analysis_df['Section_ID'] == row['Section_ID']]['Course'].iloc[0],
                'Section': analysis_df[analysis_df['Section_ID'] == row['Section_ID']]['Section'].iloc[0],
                'Teacher': row['Teachers'],
                'Period': row['Period'],
                'Room': row['Room'],
                'N_Completed': row['N_Completed'],
                'Median': row['Median'],
                'Mean': row['Mean'],
                'Pct_Struggling': row['Pct_Below_Threshold'],
                'Priority_Score': row['Priority_Score'],
                'Rank': row['Rank']
            })
    
    # Analyze each dimension
    if hasattr(config, 'dimension_groups'):
        for dimension_name, questions in config.dimension_groups.items():
            # Calculate dimension average score for each student
            dim_cols = [f'Q{q}' for q in questions if f'Q{q}' in analysis_df.columns]
            if not dim_cols:
                continue
            
            # Convert each column to numeric first, then calculate mean
            numeric_cols = analysis_df[dim_cols].apply(pd.to_numeric, errors='coerce')
            analysis_df[f'{dimension_name}_dim_avg'] = numeric_cols.mean(axis=1)
            
            section_stats = calculate_section_stats(
                analysis_df,
                'Section_ID',
                f'{dimension_name}_dim_avg',
                invert_score=False,
                threshold=2.0
            )
            
            # Store priority scores for this dimension
            dimension_priorities[dimension_name] = section_stats.set_index('Section_ID')['Priority_Score'].to_dict()
            
            for _, row in section_stats.iterrows():
                results.append({
                    'Dimension': dimension_name.replace('_', ' '),
                    'Course': analysis_df[analysis_df['Section_ID'] == row['Section_ID']]['Course'].iloc[0],
                    'Section': analysis_df[analysis_df['Section_ID'] == row['Section_ID']]['Section'].iloc[0],
                    'Teacher': row['Teachers'],
                    'Period': row['Period'],
                    'Room': row['Room'],
                    'N_Completed': row['N_Completed'],
                    'Median': row['Median'],
                    'Mean': row['Mean'],
                    'Pct_Struggling': row['Pct_Below_Threshold'],
                    'Priority_Score': row['Priority_Score'],
                    'Rank': row['Rank']
                })
    
    # Create DataFrame
    results_df = pd.DataFrame(results)
    
    # ===== ADD SET COVER RECOMMENDATIONS =====
    recommendations = calculate_set_cover_recommendations(analysis_df, config, threshold=2.0)
    results_df = add_recommendations_to_section_analysis(results_df, recommendations, config)
    
    # Sort by dimension, then by recommendation rank (prioritize recommended sections)
    results_df['_sort_key'] = results_df['Recommendation_Rank'].apply(
        lambda x: x if x != '' else 999
    )
    results_df = results_df.sort_values(['Dimension', '_sort_key', 'Rank'])
    results_df = results_df.drop('_sort_key', axis=1)
    # ===== END SET COVER CODE =====
    
    # Write to Excel
    results_df.to_excel(writer, sheet_name='Section Analysis', index=False)
    
    # Create composite priority sheet - sections with highest overall need
    create_composite_priority_sheet(writer, dimension_priorities, analysis_df)
    
    # ===== CREATE COVERAGE SUMMARY SHEET =====
    coverage_summary = create_coverage_summary(analysis_df, recommendations, config, threshold=2.0)
    if not coverage_summary.empty:
        coverage_summary.to_excel(writer, sheet_name='Coverage Summary', index=False)
    # ===== END COVERAGE SUMMARY =====


# ===== SET COVER OPTIMIZATION FUNCTIONS =====

def calculate_set_cover_recommendations(analysis_df, config, threshold=2.0):
    """
    Use weighted greedy set cover to identify top 10 sections per dimension
    that maximize coverage of at-risk students while minimizing redundancy.
    
    Parameters:
    -----------
    analysis_df : pd.DataFrame
        Student-section data with dimension scores
    config : object
        Survey configuration
    threshold : float
        Threshold for "at-risk" status (dimension avg <= threshold)
        
    Returns:
    --------
    dict
        Mapping of (dimension, section_id) -> recommendation_rank (1-10)
    """

    # exclude courses that start with the letter P or Z
    excluded_courses = ['P', 'Z']
    analysis_df = analysis_df[~analysis_df['Course'].str.startswith(tuple(excluded_courses))]

    recommendations = {}
    
    # Only process the named dimensions from config
    if not hasattr(config, 'dimension_groups'):
        return recommendations
    
    dimension_groups = config.dimension_groups
    
    for dimension_name, questions in dimension_groups.items():
        # Calculate dimension average for each student
        dim_cols = [f'Q{q}' for q in questions if f'Q{q}' in analysis_df.columns]
        if not dim_cols:
            continue
        
        # Get dimension average (should already exist from main analysis)
        dim_avg_col = f'{dimension_name}_dim_avg'
        if dim_avg_col not in analysis_df.columns:
            numeric_cols = analysis_df[dim_cols].apply(pd.to_numeric, errors='coerce')
            analysis_df[dim_avg_col] = numeric_cols.mean(axis=1)
        
        # Identify at-risk students for this dimension
        at_risk_df = analysis_df[
            pd.to_numeric(analysis_df[dim_avg_col], errors='coerce') <= threshold
        ].copy()
        
        if at_risk_df.empty:
            continue
        
        at_risk_students = set(at_risk_df['StudentID'].unique())
        
        # Build section -> student mapping
        section_students = defaultdict(set)
        section_priority = {}  # Store priority scores for tie-breaking
        
        for section_id in analysis_df['Section_ID'].unique():
            section_df = analysis_df[analysis_df['Section_ID'] == section_id]
            
            # Get at-risk students in this section
            section_at_risk = set(section_df['StudentID']) & at_risk_students
            
            if section_at_risk:
                section_students[section_id] = section_at_risk
                
                # Calculate priority score for this section-dimension combo
                # (will be used for tie-breaking)
                section_at_risk_df = section_df[
                    section_df['StudentID'].isin(section_at_risk)
                ]
                
                if len(section_at_risk_df) > 0:
                    scores = pd.to_numeric(
                        section_at_risk_df[dim_avg_col], 
                        errors='coerce'
                    ).dropna()
                    
                    if len(scores) > 0:
                        # Priority based on: number of at-risk students + severity
                        n_at_risk = len(section_at_risk)
                        avg_severity = (threshold - scores.mean()) / threshold
                        section_priority[section_id] = n_at_risk + avg_severity
                    else:
                        section_priority[section_id] = len(section_at_risk)
        
        # Run weighted greedy set cover to select top 10 sections
        selected_sections = weighted_greedy_set_cover(
            section_students,
            section_priority,
            max_sections=10
        )
        
        # Store recommendations with ranks
        for rank, section_id in enumerate(selected_sections, 1):
            recommendations[(dimension_name, section_id)] = rank
    
    return recommendations


def weighted_greedy_set_cover(section_students, section_priority, max_sections=10):
    """
    Weighted greedy set cover algorithm with partial credit for overlap.
    
    Parameters:
    -----------
    section_students : dict
        Mapping of section_id -> set of student IDs
    section_priority : dict
        Mapping of section_id -> priority score (for tie-breaking)
    max_sections : int
        Maximum number of sections to select
        
    Returns:
    --------
    list
        Ordered list of selected section IDs (ranked 1 to max_sections)
    """
    if not section_students:
        return []
    
    selected = []
    student_coverage_count = defaultdict(int)  # Track how many times each student is covered
    
    for _ in range(max_sections):
        if not section_students:
            break
        
        best_section = None
        best_score = (-1, -1)  # Initialize as tuple to match composite_score format
        
        for section_id, students in section_students.items():
            if section_id in selected:
                continue
            
            # Calculate weighted coverage score
            # Give full credit for uncovered students, partial credit for already-covered
            coverage_score = 0
            for student in students:
                if student_coverage_count[student] == 0:
                    coverage_score += 1.0  # Full credit for new coverage
                else:
                    # Diminishing returns: 1/(n+1) where n is current coverage count
                    coverage_score += 1.0 / (student_coverage_count[student] + 1)
            
            # Break ties using priority score
            composite_score = (coverage_score, section_priority.get(section_id, 0))
            
            if composite_score > best_score:
                best_score = composite_score
                best_section = section_id
        
        if best_section is None:
            break
        
        # Select this section
        selected.append(best_section)
        
        # Update coverage counts
        for student in section_students[best_section]:
            student_coverage_count[student] += 1
    
    return selected


def add_recommendations_to_section_analysis(section_analysis_df, recommendations, config):
    """
    Add recommendation ranks to the section analysis dataframe.
    
    Parameters:
    -----------
    section_analysis_df : pd.DataFrame
        Section analysis results
    recommendations : dict
        Mapping of (dimension, section_id) -> rank
    config : object
        Survey configuration (used to get dimension names)
        
    Returns:
    --------
    pd.DataFrame
        Updated dataframe with Recommendation_Rank column
    """
    section_analysis_df = section_analysis_df.copy()
    
    # Create Section_ID if it doesn't exist
    if 'Section_ID' not in section_analysis_df.columns:
        section_analysis_df['Section_ID'] = (
            section_analysis_df['Course'] + ' - ' + section_analysis_df['Section']
        )
    
    # Build dimension mapping from config
    dimension_map = {}
    if hasattr(config, 'dimension_groups'):
        for internal_name in config.dimension_groups.keys():
            display_name = internal_name.replace('_', ' ')
            dimension_map[display_name] = internal_name
    
    # Add recommendation rank
    def get_recommendation_rank(row):
        dimension = row['Dimension']
        section_id = row.get('Section_ID', '')
        
        # Convert dimension display name back to internal name
        internal_dimension = dimension_map.get(dimension, dimension)
        rank = recommendations.get((internal_dimension, section_id), None)
        
        return rank if rank is not None else ''
    
    section_analysis_df['Recommendation_Rank'] = section_analysis_df.apply(
        get_recommendation_rank, 
        axis=1
    )
    
    return section_analysis_df


def create_coverage_summary(analysis_df, recommendations, config, threshold=2.0):
    """
    Create a summary showing coverage statistics for the recommended sections.
    
    Parameters:
    -----------
    analysis_df : pd.DataFrame
        Student-section data
    recommendations : dict
        Recommendation mappings
    config : object
        Survey configuration
    threshold : float
        At-risk threshold
        
    Returns:
    --------
    pd.DataFrame
        Summary statistics per dimension
    """
    summary_data = []
    
    if not hasattr(config, 'dimension_groups'):
        return pd.DataFrame()
    
    for dimension_name, questions in config.dimension_groups.items():
        dim_cols = [f'Q{q}' for q in questions if f'Q{q}' in analysis_df.columns]
        if not dim_cols:
            continue
        
        dim_avg_col = f'{dimension_name}_dim_avg'
        if dim_avg_col not in analysis_df.columns:
            continue
        
        # Get at-risk students
        at_risk_df = analysis_df[
            pd.to_numeric(analysis_df[dim_avg_col], errors='coerce') <= threshold
        ]
        total_at_risk = len(at_risk_df['StudentID'].unique())
        
        if total_at_risk == 0:
            continue
        
        # Get recommended sections for this dimension
        recommended_sections = [
            section_id for (dim, section_id), rank in recommendations.items()
            if dim == dimension_name
        ]
        
        if not recommended_sections:
            continue
        
        # Calculate coverage
        covered_students = set()
        duplicate_coverage_count = 0
        
        for section_id in recommended_sections:
            section_students = set(
                analysis_df[
                    (analysis_df['Section_ID'] == section_id) &
                    (analysis_df['StudentID'].isin(at_risk_df['StudentID']))
                ]['StudentID']
            )
            
            # Track duplicates
            already_covered = covered_students & section_students
            duplicate_coverage_count += len(already_covered)
            
            covered_students.update(section_students)
        
        coverage_pct = (len(covered_students) / total_at_risk * 100) if total_at_risk > 0 else 0
        avg_coverage_per_student = (len(covered_students) + duplicate_coverage_count) / len(covered_students) if len(covered_students) > 0 else 0
        
        summary_data.append({
            'Dimension': dimension_name.replace('_', ' '),
            'Total_At_Risk_Students': total_at_risk,
            'Students_Covered': len(covered_students),
            'Coverage_Pct': round(coverage_pct, 1),
            'Recommended_Sections': len(recommended_sections),
            'Avg_Coverage_Per_Student': round(avg_coverage_per_student, 2),
            'Total_Student_Interventions': len(covered_students) + duplicate_coverage_count
        })
    
    return pd.DataFrame(summary_data)

# ===== END SET COVER OPTIMIZATION FUNCTIONS =====


def create_composite_priority_sheet(writer, dimension_priorities, analysis_df):
    """
    Create a sheet showing sections ranked by composite priority across ALL dimensions.
    This identifies the highest-leverage classrooms for Tier 2 interventions.
    
    Parameters:
    -----------
    writer : pd.ExcelWriter
        Excel writer object
    dimension_priorities : dict
        Dictionary mapping dimension_name -> {section_id: priority_score}
    analysis_df : pd.DataFrame
        Analysis dataframe with section info
    """
    
    if not dimension_priorities:
        return
    
    # Get all unique sections
    all_sections = set()
    for dim_scores in dimension_priorities.values():
        all_sections.update(dim_scores.keys())
    
    composite_results = []
    
    for section_id in all_sections:
        # Calculate composite priority score (average across all dimensions)
        dim_scores = []
        dim_details = {}
        
        for dim_name, scores in dimension_priorities.items():
            if section_id in scores:
                dim_scores.append(scores[section_id])
                dim_details[f'{dim_name}_Priority'] = round(scores[section_id], 3)
        
        if not dim_scores:
            continue
        
        # Composite score is the average of dimension priority scores
        composite_score = sum(dim_scores) / len(dim_scores)
        
        # Count how many dimensions this section is high-priority in (top 20%)
        high_priority_dims = sum(1 for score in dim_scores if score >= 0.6)
        
        # Get section details
        section_df = analysis_df[analysis_df['Section_ID'] == section_id]
        if section_df.empty:
            continue
        
        course = section_df['Course'].iloc[0]
        section = section_df['Section'].iloc[0]
        
        teachers = []
        if 'Teacher1' in section_df.columns and section_df['Teacher1'].iloc[0]:
            teachers.append(str(section_df['Teacher1'].iloc[0]))
        if 'Teacher2' in section_df.columns and section_df['Teacher2'].iloc[0]:
            teachers.append(str(section_df['Teacher2'].iloc[0]))
        teacher_str = ', '.join(teachers) if teachers else ''
        
        period = str(section_df['Period'].iloc[0]) if 'Period' in section_df.columns else ''
        room = str(section_df['Room'].iloc[0]) if 'Room' in section_df.columns else ''
        
        n_students = len(section_df)
        
        result = {
            'Course': course,
            'Section': section,
            'Teacher': teacher_str,
            'Period': period,
            'Room': room,
            'N_Students': n_students,
            'Composite_Priority': round(composite_score, 3),
            'High_Priority_Dimensions': high_priority_dims,
            'Num_Dimensions_Analyzed': len(dim_scores)
        }
        
        # Add individual dimension scores
        result.update(dim_details)
        
        composite_results.append(result)
    
    # Create DataFrame and sort by composite priority
    composite_df = pd.DataFrame(composite_results)
    composite_df = composite_df.sort_values('Composite_Priority', ascending=False)
    composite_df['Overall_Rank'] = range(1, len(composite_df) + 1)
    
    # Reorder columns to put rank first
    cols = ['Overall_Rank', 'Course', 'Section', 'Teacher', 'Period', 'Room', 
            'N_Students', 'Composite_Priority', 'High_Priority_Dimensions', 'Num_Dimensions_Analyzed']
    
    # Add dimension priority columns at the end
    dim_cols = [col for col in composite_df.columns if col.endswith('_Priority')]
    cols.extend(sorted(dim_cols))
    
    composite_df = composite_df[cols]
    
    # Write to Excel
    composite_df.to_excel(writer, sheet_name='Composite Priority', index=False)


def calculate_section_stats(df, section_col, score_col, invert_score=False, threshold=2.0):
    """
    Calculate statistics for each section.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Data with section and score columns
    section_col : str
        Column name for section identifier
    score_col : str
        Column name for score to analyze
    invert_score : bool
        If True, lower scores are better (for Weighted_Risk_Score)
    threshold : float
        Threshold for determining "struggling" students
        
    Returns:
    --------
    pd.DataFrame
        Section statistics sorted by priority
    """
    results = []
    
    for section in df[section_col].unique():
        section_df = df[df[section_col] == section].copy()
        
        # Get teacher, period, room info
        teachers = []
        if 'Teacher1' in section_df.columns and section_df['Teacher1'].iloc[0]:
            teachers.append(str(section_df['Teacher1'].iloc[0]))
        if 'Teacher2' in section_df.columns and section_df['Teacher2'].iloc[0]:
            teachers.append(str(section_df['Teacher2'].iloc[0]))
        teacher_str = ', '.join(teachers) if teachers else ''
        
        period = str(section_df['Period'].iloc[0]) if 'Period' in section_df.columns else ''
        room = str(section_df['Room'].iloc[0]) if 'Room' in section_df.columns else ''
        
        # Calculate statistics
        scores = pd.to_numeric(section_df[score_col], errors='coerce').dropna()
        n = len(scores)
        
        if n == 0:
            continue
        
        median = scores.median()
        mean = scores.mean()
        
        # Calculate % struggling
        if invert_score:
            # For Weighted_Risk_Score: lower is better, so struggling = bottom 25%
            threshold_val = scores.quantile(0.25)
            pct_struggling = (scores <= threshold_val).sum() / n * 100
        else:
            # For dimension scores: higher is better, so struggling = score <= threshold
            pct_struggling = (scores <= threshold).sum() / n * 100
        
        # Calculate priority score (higher = higher priority for intervention)
        # Weight: 50% pct struggling, 30% median (inverted), 20% sample size (normalized)
        if invert_score:
            median_component = (median - scores.min()) / (scores.max() - scores.min() + 0.001)
        else:
            median_component = 1 - ((median - scores.min()) / (scores.max() - scores.min() + 0.001))
        
        sample_size_component = min(n / 30, 1.0)  # Normalize to max of 30 students
        
        priority_score = (
            0.5 * (pct_struggling / 100) +
            0.3 * median_component +
            0.2 * sample_size_component
        )
        
        results.append({
            'Section_ID': section,
            'Teachers': teacher_str,
            'Period': period,
            'Room': room,
            'N_Completed': n,
            'Median': round(median, 2),
            'Mean': round(mean, 2),
            'Pct_Below_Threshold': round(pct_struggling, 1),
            'Priority_Score': round(priority_score, 3)
        })
    
    results_df = pd.DataFrame(results)
    results_df = results_df.sort_values('Priority_Score', ascending=False)
    results_df['Rank'] = range(1, len(results_df) + 1)
    
    return results_df


def format_workbook(wb):
    """Apply formatting to all sheets in workbook."""
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    tier3_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    tier2_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    tier1_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    high_priority_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format headers
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add conditional formatting for tier sheets
        if sheet_name in ['Tier 3', 'Tier 2', 'Tier 1']:
            format_tier_sheet(ws, sheet_name)
        
        # Format Section Analysis sheet
        if sheet_name == 'Section Analysis':
            format_section_analysis_sheet(ws)
        
        # Format Composite Priority sheet
        if sheet_name == 'Composite Priority':
            format_composite_priority_sheet(ws)
        
        # Format Coverage Summary sheet
        if sheet_name == 'Coverage Summary':
            format_coverage_summary_sheet(ws)


def format_tier_sheet(ws, tier_name):
    """Apply tier-specific formatting."""
    if tier_name == 'Tier 3':
        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif tier_name == 'Tier 2':
        fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Highlight rank column
    if ws.max_row > 1:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=1).font = Font(bold=True)


def format_section_analysis_sheet(ws):
    """Format section analysis sheet with highlighting for recommended sections."""
    
    if ws.max_row < 2:
        return
    
    # Find columns
    rank_col = None
    rec_rank_col = None
    
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Rank':
            rank_col = idx
        elif cell.value == 'Recommendation_Rank':
            rec_rank_col = idx
    
    if rank_col is None:
        return
    
    # Define fills
    high_priority_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    medium_priority_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    recommended_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    top_recommended_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    current_dimension = None
    for row in range(2, ws.max_row + 1):
        dimension_cell = ws.cell(row=row, column=1)
        rank_cell = ws.cell(row=row, column=rank_col)
        
        # Track dimension changes
        if dimension_cell.value != current_dimension:
            current_dimension = dimension_cell.value
        
        # Highlight based on recommendation rank (if present)
        if rec_rank_col:
            rec_rank_cell = ws.cell(row=row, column=rec_rank_col)
            
            if rec_rank_cell.value and rec_rank_cell.value != '':
                try:
                    rec_rank = int(rec_rank_cell.value)
                    
                    if rec_rank <= 3:
                        # Top 3 recommendations - bright green
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = top_recommended_fill
                        rec_rank_cell.font = Font(bold=True, size=12)
                    elif rec_rank <= 10:
                        # Recommendations 4-10 - light green
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = recommended_fill
                        rec_rank_cell.font = Font(bold=True)
                except (ValueError, TypeError):
                    pass
                continue
        
        # If no recommendation rank, fall back to old ranking system
        if rank_cell.value and isinstance(rank_cell.value, (int, float)):
            if rank_cell.value <= 3:
                for col in range(1, rank_col + 1):
                    ws.cell(row=row, column=col).fill = high_priority_fill
            elif rank_cell.value <= 5:
                for col in range(1, rank_col + 1):
                    ws.cell(row=row, column=col).fill = medium_priority_fill
    
    # Format percentage column
    pct_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Pct_Struggling':
            pct_col = idx
            break
    
    if pct_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=pct_col)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '0.0'


def format_composite_priority_sheet(ws):
    """Format composite priority sheet with highlighting for highest-leverage classrooms."""
    
    if ws.max_row < 2:
        return
    
    # Find the Overall_Rank column
    rank_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Overall_Rank':
            rank_col = idx
            break
    
    if rank_col is None:
        return
    
    # Highlight top sections
    high_priority_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    medium_priority_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_medium_priority_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    for row in range(2, ws.max_row + 1):
        rank_cell = ws.cell(row=row, column=rank_col)
        
        # Highlight based on overall rank
        if rank_cell.value and isinstance(rank_cell.value, (int, float)):
            if rank_cell.value <= 5:
                # Top 5: High priority (RED)
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = high_priority_fill
                    if col == rank_col:
                        ws.cell(row=row, column=col).font = Font(bold=True, size=12)
            elif rank_cell.value <= 10:
                # 6-10: Medium-high priority (ORANGE)
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = medium_priority_fill
            elif rank_cell.value <= 15:
                # 11-15: Medium priority (YELLOW)
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = low_medium_priority_fill
    
    # Format Composite_Priority column as decimal
    comp_priority_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Composite_Priority':
            comp_priority_col = idx
            break
    
    if comp_priority_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=comp_priority_col)
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '0.000'


def format_coverage_summary_sheet(ws):
    """Format coverage summary sheet."""
    
    if ws.max_row < 2:
        return
    
    # Format percentage column
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Coverage_Pct':
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = '0.0"%"'
        elif cell.value == 'Avg_Coverage_Per_Student':
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = '0.00'