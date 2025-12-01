"""
Main orchestrator for Weekly Period Attendance Analysis Report
Generates comprehensive Excel workbook with multiple sheets for different stakeholders
"""

import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime, timedelta
from flask import session
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference

from app.scripts.attendance.jupiter.process import process_local_file as process_jupiter

# Import helper modules
from . import attendance_data_prep as data_prep
from . import attendance_analysis_utils as analysis
from . import attendance_formatting_utils as formatting


"""
Updated main() function for the orchestrator
Replace the counselor and attendance teacher sheet creation sections
"""

def main(form, request):
    """
    Main function to generate weekly attendance analysis report
    
    Args:
        form: Flask form with week_of selection
        request: Flask request object
    
    Returns:
        tuple: (BytesIO buffer, filename)
    """
    week_number = form.week_of.data
    
    # Load all semester attendance data
    print(f"Loading attendance data for week {week_number}...")
    attendance_df = process_jupiter(week_number=None)  # Load all weeks
    
    # Add attendance teacher assignment
    attendance_df['AttendanceTeacher'] = attendance_df.apply(
        data_prep.return_attendance_teacher, axis=1
    )
    
    # Ensure Date is datetime and add week_id (ISO week number)
    attendance_df['Date'] = pd.to_datetime(attendance_df['Date'])
    attendance_df['week_id'] = attendance_df['Date'].dt.isocalendar().week
    
    # Get date range for selected week
    selected_week_df = attendance_df[attendance_df['week_number'] == int(week_number)].copy()
    
    if len(selected_week_df) == 0:
        raise ValueError(f"No data found for week {week_number}")
    
    # Determine the max date in the selected week
    selected_week_dates = sorted(selected_week_df['Date'].unique())
    max_selected_date = max(selected_week_dates)
    
    # ✓ FILTER OUT FUTURE DATA: Only keep attendance data up through the selected week
    print(f"Filtering out data after {max_selected_date}...")
    attendance_df = attendance_df[attendance_df['Date'] <= max_selected_date].copy()
    print(f"Remaining records after filtering: {len(attendance_df)}")
    
    # Now recalculate selected_week_df after filtering
    selected_week_df = attendance_df[attendance_df['week_number'] == int(week_number)].copy()
    
    if len(selected_week_df) == 0:
        raise ValueError(f"No data found for week {week_number} after filtering future dates")
    
    selected_week_dates = sorted(selected_week_df['Date'].unique())
    start_date = min(selected_week_dates)
    end_date = max(selected_week_dates)
    
    if start_date == end_date:
        date_of_report = str(start_date)[:10]
    else:
        date_of_report = f"{str(start_date)[:10]} to {str(end_date)[:10]}"
    
    print(f"Analyzing week {week_number}: {date_of_report}")
    print(f"Total attendance records through this week: {len(attendance_df)}")
    
    # Calculate weekly statistics and trends
    print("Calculating weekly statistics and semester trends...")
    weekly_stats = data_prep.aggregate_weekly_attendance(attendance_df)
    weekly_stats = data_prep.calculate_semester_trends(weekly_stats)
    
    # Debug: Check if week_id exists in both dataframes
    print(f"Selected week_df week_ids: {selected_week_df['week_id'].unique()}")
    print(f"Weekly_stats week_ids: {weekly_stats['week_id'].unique()}")
    print(f"Weekly_stats shape: {weekly_stats.shape}")
    
    # Identify selected week_id with better error handling
    matching_weeks = weekly_stats[
        weekly_stats['week_id'].isin(
            selected_week_df['week_id'].unique()
        )
    ]
    
    if len(matching_weeks) == 0:
        # Fallback: use the week_id from selected_week_df directly
        selected_week_id = selected_week_df['week_id'].iloc[0]
        print(f"Warning: No matching week_id found in weekly_stats. Using {selected_week_id} from selected_week_df")
        
        # If weekly_stats is completely empty, that's a bigger problem
        if len(weekly_stats) == 0:
            raise ValueError(
                f"aggregate_weekly_attendance returned no data. "
                f"Check that attendance_df has valid data with proper grouping columns."
            )
    else:
        selected_week_id = matching_weeks['week_id'].iloc[0]
        print(f"Selected week_id: {selected_week_id}")
    
    # Calculate summary statistics
    print("Calculating summary statistics...")
    summary_stats = analysis.calculate_weekly_summary_stats(
        attendance_df, weekly_stats, selected_week_dates, selected_week_id
    )
    
    # Identify Tier 3 students
    print("Identifying Tier 3 students...")
    tier3_students = data_prep.identify_tier3_students(
        weekly_stats, selected_week_id, top_pct=0.05
    )
    
    # Add attendance teacher to tier3
    tier3_students['AttendanceTeacher'] = tier3_students.apply(
        data_prep.return_attendance_teacher, axis=1
    )
    
    # Calculate most improved students
    print("Calculating most improved students...")
    most_improved_by_section = analysis.calculate_most_improved_by_section(
        weekly_stats, selected_week_id
    )
    most_improved_by_counselor = analysis.calculate_most_improved_by_counselor(
        weekly_stats, selected_week_id, top_n=10
    )
    
    # Get unique lists of stakeholders
    teachers = sorted(attendance_df['Teacher'].unique())
    counselors = sorted(attendance_df['Counselor'].unique())
    attendance_teachers = sorted(attendance_df['AttendanceTeacher'].unique())
    
    print(f"Creating workbook for {len(teachers)} teachers, {len(counselors)} counselors, {len(attendance_teachers)} attendance teachers")
    
    # Create Excel workbook
    output = BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Sheet 1: Weekly Summary
    print("Creating Weekly Summary sheet...")
    create_weekly_summary_sheet(
        wb, summary_stats, date_of_report, weekly_stats, selected_week_id
    )
    
    # Sheet 2: Tier 3 Watch List
    print("Creating Tier 3 Watch List sheet...")
    create_tier3_sheet(wb, tier3_students, date_of_report)
    
    # Sheet 3: Most Improved
    print("Creating Most Improved sheet...")
    create_most_improved_sheet(
        wb, most_improved_by_section, most_improved_by_counselor, date_of_report
    )
    
    # Counselor sheets (UPDATED - now passes tier3_students)
    print("Creating counselor sheets...")
    for counselor in counselors:
        create_counselor_sheet(
            wb, counselor, attendance_df, weekly_stats, selected_week_dates,
            selected_week_id, most_improved_by_counselor, tier3_students  # ← Added tier3_students
        )
    
    # Attendance teacher sheets (UPDATED - now passes tier3_students)
    print("Creating attendance teacher sheets...")
    for att_teacher in attendance_teachers:
        create_attendance_teacher_sheet(
            wb, att_teacher, attendance_df, weekly_stats, selected_week_dates,
            selected_week_id, tier3_students  # ← Added tier3_students
        )
    
    # Individual teacher sheets
    print("Creating individual teacher sheets...")
    for teacher in teachers:
        create_teacher_sheet(
            wb, teacher, attendance_df, weekly_stats, selected_week_dates,
            selected_week_id, most_improved_by_section
        )
    
    # Save workbook
    print("Saving workbook...")
    wb.save(output)
    output.seek(0)
    
    filename = f"JupiterAttendanceAnalysisReport-{date_of_report}.xlsx"
    print(f"Report generated: {filename}")
    
    return output, filename


def create_weekly_summary_sheet(wb, summary_stats, date_of_report, weekly_stats, selected_week_id):
    """Create the weekly summary overview sheet"""
    ws = wb.create_sheet("Weekly Summary")
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:F{current_row}')
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = f"Attendance Analysis Summary - {date_of_report}"
    title_cell.font = openpyxl.styles.Font(bold=True, size=14)
    title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    current_row += 2
    
    # Overall statistics
    current_row = formatting.create_summary_box(
        ws, current_row, 1, "Overall Attendance Rates",
        {
            'This Week': formatting.format_percentage(summary_stats['week_attendance_rate']),
            'Semester Average': formatting.format_percentage(summary_stats['semester_attendance_rate']),
            'Difference': formatting.format_percentage(
                summary_stats['week_attendance_rate'] - summary_stats['semester_attendance_rate']
            ),
        }
    )
    current_row += 1
    
    # By year in HS
    current_row = formatting.create_summary_box(
        ws, current_row, 1, "Attendance by Year",
        {
            f"Year {year} - This Week": formatting.format_percentage(
                summary_stats['week_by_year'].get(year, 0)
            )
            for year in sorted(summary_stats['week_by_year'].keys())
        }
    )
    current_row += 2
    
    # Issues summary
    current_row = formatting.create_summary_box(
        ws, current_row, 4, "Issues This Week",
        {
            'Potential Cuts': summary_stats['week_cut_count'],
            'Late to School': summary_stats['week_late_count'],
            'Chronic Absent Students': summary_stats.get('chronic_absent_count', 0),
        }
    )
    
    # Add trend chart
    current_row += 2
    create_trend_chart(ws, weekly_stats, selected_week_id, current_row)
    
    # Auto-adjust columns
    formatting.auto_adjust_column_width(ws)


def create_tier3_sheet(wb, tier3_df, date_of_report):
    """Create Tier 3 watch list sheet"""
    ws = wb.create_sheet("Tier 3 Watch List")
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:K{current_row}')
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = f"Tier 3 Students Requiring Intervention - {date_of_report}"
    title_cell.font = openpyxl.styles.Font(bold=True, size=14, color='FF0000')
    current_row += 2
    
    if len(tier3_df) == 0:
        ws.cell(row=current_row, column=1).value = "No Tier 3 students identified this week."
        return
    
    # ✓ FIXED: Changed 'Teachers' to 'Teacher' to match the actual column name
    output_df = tier3_df[[
        'year_in_hs', 'LastName', 'FirstName', 'Counselor', 'AttendanceTeacher',
        'Teacher',  # ✓ FIXED: was 'Teachers'
        'semester_attendance_rate', 'attendance_rate_smooth',
        'attendance_trend', 'cut_count', 'late_count'
    ]].copy()
    
    output_df.columns = [
        'Year', 'Last Name', 'First Name', 'Counselor', 'Attendance Teacher',
        'Teachers',  # Display name is plural for clarity
        'Semester Rate', 'Recent Rate', 'Trend',
        'Cuts', 'Lates'
    ]
    
    # Format percentages
    output_df['Semester Rate'] = output_df['Semester Rate'].apply(formatting.format_percentage)
    output_df['Recent Rate'] = output_df['Recent Rate'].apply(formatting.format_percentage)
    output_df['Trend'] = output_df['Trend'].apply(formatting.format_trend)
    
    # Write to sheet
    for r_idx, row in enumerate(dataframe_to_rows(output_df, index=False, header=True), current_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == current_row:  # Header row
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color=formatting.COLORS['header'],
                    end_color=formatting.COLORS['header'],
                    fill_type='solid'
                )
            else:
                # Color code by severity
                semester_rate = tier3_df.iloc[r_idx - current_row - 1]['semester_attendance_rate']
                if semester_rate < 0.8:
                    fill_color = formatting.COLORS['chronic_absent']
                elif semester_rate < 0.9:
                    fill_color = formatting.COLORS['approaching_chronic']
                else:
                    fill_color = formatting.COLORS['declining_trend']
                
                cell.fill = openpyxl.styles.PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type='solid'
                )
    
    formatting.auto_adjust_column_width(ws)


def create_most_improved_sheet(wb, by_section_df, by_counselor_df, date_of_report):
    """Create most improved students sheet"""
    ws = wb.create_sheet("Most Improved")
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:H{current_row}')
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = f"Most Improved Attendance - {date_of_report}"
    title_cell.font = openpyxl.styles.Font(bold=True, size=14, color='008000')
    current_row += 2
    
    # Section: By course section
    current_row = formatting.create_section_header(
        ws, current_row, "Most Improved by Course Section", 8
    )
    
    if len(by_section_df) > 0:
        section_output = by_section_df[[
            'LastName', 'FirstName', 'Course', 'Section', 'Teacher',
            'BaselineRate', 'RecentRate', 'Improvement'
        ]].copy()
        
        section_output['BaselineRate'] = section_output['BaselineRate'].apply(formatting.format_percentage)
        section_output['RecentRate'] = section_output['RecentRate'].apply(formatting.format_percentage)
        section_output['Improvement'] = section_output['Improvement'].apply(formatting.format_percentage)
        
        section_output.columns = [
            'Last Name', 'First Name', 'Course', 'Section', 'Teacher',
            'Baseline', 'Recent', 'Improvement'
        ]
        
        for r_idx, row in enumerate(dataframe_to_rows(section_output, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(section_output.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(section_output.columns), 'most_improved')
        
        current_row += len(section_output) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No improved students by section"
        current_row += 2
    
    # Section: Top 10 by counselor
    current_row = formatting.create_section_header(
        ws, current_row, "Top 10 Most Improved by Counselor", 8
    )
    
    if len(by_counselor_df) > 0:
        counselor_output = by_counselor_df[[
            'LastName', 'FirstName', 'Counselor', 'year_in_hs',
            'AvgAttendanceRate', 'AvgTrend', 'ImprovementScore', 'Rank'
        ]].copy()
        
        counselor_output['AvgAttendanceRate'] = counselor_output['AvgAttendanceRate'].apply(formatting.format_percentage)
        counselor_output['AvgTrend'] = counselor_output['AvgTrend'].apply(formatting.format_trend)
        
        counselor_output.columns = [
            'Last Name', 'First Name', 'Counselor', 'Year',
            'Avg Rate', 'Trend', 'Score', 'Rank'
        ]
        
        for r_idx, row in enumerate(dataframe_to_rows(counselor_output, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(counselor_output.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(counselor_output.columns), 'most_improved')
        
        current_row += len(counselor_output) + 2
    
    formatting.auto_adjust_column_width(ws)


def create_teacher_sheet(wb, teacher, attendance_df, weekly_stats, selected_week_dates, selected_week_id, most_improved_df):
    """Create individual teacher sheet with attendance analysis"""
    # Sanitize sheet name (max 31 chars, no special chars)
    sheet_name = teacher[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('[', '').replace(']', '')
    ws = wb.create_sheet(sheet_name)
    
    current_row = 1
    
    # Title
    ws.merge_cells(f'A{current_row}:J{current_row}')
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = f"Attendance Report - {teacher}"
    title_cell.font = openpyxl.styles.Font(bold=True, size=14)
    current_row += 2
    
    # Attendance completion grid
    current_row = formatting.create_section_header(ws, current_row, "Attendance Submission Status", 10)
    
    completion_grid = analysis.create_attendance_completion_grid(
        attendance_df, teacher, selected_week_dates
    )
    
    if len(completion_grid) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(completion_grid, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(completion_grid.columns))
                else:
                    # Color missing attendance
                    if value == '✗':
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color=formatting.COLORS['missing_data'],
                            end_color=formatting.COLORS['missing_data'],
                            fill_type='solid'
                        )
        current_row += len(completion_grid) + 2
    
    # Students with potential cuts
    current_row = formatting.create_section_header(ws, current_row, "Potential Cuts - Requires Follow-Up", 10)
    
    cuts_df = analysis.get_students_with_cuts(attendance_df, teacher, selected_week_dates)
    
    if len(cuts_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(cuts_df, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(cuts_df.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(cuts_df.columns), 'cuts')
        current_row += len(cuts_df) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No potential cuts this week"
        current_row += 2
    
    # Students late to school
    current_row = formatting.create_section_header(ws, current_row, "Students Late to School", 8)
    
    lates_df = analysis.get_students_late_to_school(attendance_df, teacher, selected_week_dates)
    
    if len(lates_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(lates_df, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(lates_df.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(lates_df.columns), 'late_to_school')
        current_row += len(lates_df) + 2
    
    # Attendance errors
    current_row = formatting.create_section_header(ws, current_row, "Potential Attendance Errors", 8)
    
    errors_df = analysis.get_attendance_errors(attendance_df, teacher, selected_week_dates)
    
    if len(errors_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(errors_df, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(errors_df.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(errors_df.columns), 'attendance_error')
        current_row += len(errors_df) + 2
    
    # Students absent all week
    current_row = formatting.create_section_header(ws, current_row, "Students Absent All Week", 8)
    
    absent_df = analysis.get_students_absent_all_week(attendance_df, teacher, selected_week_dates)
    
    if len(absent_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(absent_df, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(absent_df.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(absent_df.columns), 'chronic_absent')
        current_row += len(absent_df) + 2
    
    # Most improved in teacher's sections
    teacher_improved = most_improved_df[most_improved_df['Teacher'] == teacher]
    if len(teacher_improved) > 0:
        current_row = formatting.create_section_header(ws, current_row, "Most Improved Students", 8)
        
        improved_output = teacher_improved[[
            'LastName', 'FirstName', 'Course', 'Section', 'Improvement'
        ]].copy()
        improved_output['Improvement'] = improved_output['Improvement'].apply(formatting.format_percentage)
        
        for r_idx, row in enumerate(dataframe_to_rows(improved_output, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(improved_output.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(improved_output.columns), 'most_improved')
    
    formatting.auto_adjust_column_width(ws)

"""
Enhanced sheet creation functions for counselor and attendance teacher reports
Replace the existing create_counselor_sheet and create_attendance_teacher_sheet functions
"""

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from . import attendance_formatting_utils as formatting


def create_counselor_sheet(wb, counselor, attendance_df, weekly_stats, selected_week_dates, 
                           selected_week_id, most_improved_df, tier3_df):
    """
    Enhanced counselor sheet with comprehensive student analysis
    
    Args:
        wb: Workbook object
        counselor: Counselor name
        attendance_df: Full attendance dataframe
        weekly_stats: Weekly aggregated stats
        selected_week_dates: Dates in selected week
        selected_week_id: Week ID
        most_improved_df: Most improved students dataframe
        tier3_df: Tier 3 students dataframe
    """
    from . import attendance_analysis_utils_enhanced as enhanced
    
    sheet_name = counselor[:31].replace('/', '-').replace('\\', '-')
    ws = wb.create_sheet(f"C-{sheet_name}")
    
    # Filter for this counselor
    student_filter = {'Counselor': counselor}
    counselor_data = attendance_df[attendance_df['Counselor'] == counselor].copy()
    
    current_row = 1
    
    # ===== TITLE =====
    ws.merge_cells(f'A{current_row}:J{current_row}')
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = f"Counselor Report - {counselor}"
    title_cell.font = openpyxl.styles.Font(bold=True, size=14)
    current_row += 2
    
    # ===== SUMMARY STATS =====
    counselor_students = counselor_data['StudentID'].unique()
    week_data = counselor_data[counselor_data['Date'].isin(selected_week_dates)]
    
    total_periods = len(week_data)
    present_periods = len(week_data[week_data['Type'].isin(['present', 'tardy'])])
    attendance_rate = present_periods / total_periods if total_periods > 0 else 0
    
    current_row = formatting.create_summary_box(
        ws, current_row, 1, f"Caseload Summary ({len(counselor_students)} students)",
        {
            'Attendance Rate This Week': formatting.format_percentage(attendance_rate),
            'Total Absences': len(week_data[week_data['Type'] == 'unexcused']),
            'Potential Cuts': len(week_data[week_data['cutting?'] == True]),
            'Late to School': len(week_data[week_data['late_to_school?'] == True]),
        }
    )
    current_row += 2
    
    # ===== SECTION 1: TIER 3 STUDENTS =====
    current_row = formatting.create_section_header(
        ws, current_row, "⚠️ TIER 3 STUDENTS - Requiring Immediate Intervention", 10
    )
    
    tier3_caseload = enhanced.get_tier3_students_for_caseload(tier3_df, student_filter)
    
    if len(tier3_caseload) > 0:
        # Format percentages
        tier3_display = tier3_caseload.copy()
        tier3_display['SemesterRate'] = tier3_display['SemesterRate'].apply(formatting.format_percentage)
        tier3_display['RecentRate'] = tier3_display['RecentRate'].apply(formatting.format_percentage)
        tier3_display['Trend'] = tier3_display['Trend'].apply(formatting.format_trend)
        
        for r_idx, row in enumerate(dataframe_to_rows(tier3_display, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(tier3_display.columns))
                else:
                    # Color by severity
                    semester_rate = tier3_caseload.iloc[r_idx - current_row - 1]['SemesterRate']
                    if semester_rate < 0.8:
                        color = 'chronic_absent'
                    elif semester_rate < 0.9:
                        color = 'approaching_chronic'
                    else:
                        color = 'declining_trend'
                    formatting.apply_data_row_style(ws, r_idx, len(tier3_display.columns), color)
        
        current_row += len(tier3_display) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No Tier 3 students in caseload"
        current_row += 2
    
    # ===== SECTION 2: STUDENTS WITH CUTS =====
    current_row = formatting.create_section_header(
        ws, current_row, "🚫 STUDENTS WITH CUTS THIS WEEK", 10
    )
    
    cuts_df = enhanced.get_cuts_by_period_detailed(
        attendance_df, selected_week_dates, student_filter
    )
    
    if len(cuts_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(cuts_df, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(cuts_df.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(cuts_df.columns), 'cuts')
                    # Wrap text for period breakdown column
                    if c_idx == len(cuts_df.columns):  # Last column (PeriodBreakdown)
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        
        current_row += len(cuts_df) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No cuts this week"
        current_row += 2
    
    # ===== SECTION 3: STUDENTS LATE TO SCHOOL =====
    current_row = formatting.create_section_header(
        ws, current_row, "⏰ STUDENTS LATE TO SCHOOL THIS WEEK", 14
    )
    
    late_df = enhanced.get_late_to_school_detailed(
        attendance_df, selected_week_dates, student_filter
    )
    
    if len(late_df) > 0:
        # Format for display
        late_display = late_df.copy()
        
        for r_idx, row in enumerate(dataframe_to_rows(late_display, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(late_display.columns))
                else:
                    # Color by trend
                    trend = late_df.iloc[r_idx - current_row - 1]['Trend']
                    if trend == 'Worsening':
                        color = 'late_to_school'
                    elif trend == 'Improving':
                        color = 'most_improved'
                    else:
                        color = None
                    formatting.apply_data_row_style(ws, r_idx, len(late_display.columns), color)
                    
                    # Wrap text for dates column
                    if c_idx == len(late_display.columns):  # Last column
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        
        current_row += len(late_display) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No students late to school this week"
        current_row += 2
    
    # ===== SECTION 4: STUDENTS ABSENT ALL WEEK =====
    current_row = formatting.create_section_header(
        ws, current_row, "❌ STUDENTS ABSENT ALL WEEK", 10
    )
    
    absent_df = enhanced.get_absent_all_week_with_trends(
        attendance_df, weekly_stats, selected_week_dates, 
        selected_week_id, student_filter
    )
    
    if len(absent_df) > 0:
        # Format percentages
        absent_display = absent_df.copy()
        absent_display['SemesterRate'] = absent_display['SemesterRate'].apply(formatting.format_percentage)
        absent_display['RecentRate'] = absent_display['RecentRate'].apply(formatting.format_percentage)
        
        for r_idx, row in enumerate(dataframe_to_rows(absent_display, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(absent_display.columns))
                else:
                    # Color by trend
                    trend = absent_df.iloc[r_idx - current_row - 1]['Trend']
                    if trend == 'Declining':
                        color = 'chronic_absent'
                    elif trend == 'Improving':
                        color = 'approaching_chronic'
                    else:
                        color = 'chronic_absent'
                    formatting.apply_data_row_style(ws, r_idx, len(absent_display.columns), color)
        
        current_row += len(absent_display) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No students absent all week"
        current_row += 2
    
    # ===== SECTION 5: MOST IMPROVED =====
    counselor_improved = most_improved_df[most_improved_df['Counselor'] == counselor]
    if len(counselor_improved) > 0:
        current_row = formatting.create_section_header(
            ws, current_row, "⭐ TOP 10 MOST IMPROVED STUDENTS", 8
        )
        
        improved_output = counselor_improved[[
            'LastName', 'FirstName', 'year_in_hs', 'AvgAttendanceRate', 
            'AvgTrend', 'Rank'
        ]].copy()
        improved_output['AvgAttendanceRate'] = improved_output['AvgAttendanceRate'].apply(formatting.format_percentage)
        improved_output['AvgTrend'] = improved_output['AvgTrend'].apply(formatting.format_trend)
        
        improved_output.columns = ['Last Name', 'First Name', 'Year', 'Avg Rate', 'Trend', 'Rank']
        
        for r_idx, row in enumerate(dataframe_to_rows(improved_output, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(improved_output.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(improved_output.columns), 'most_improved')
        
        current_row += len(improved_output) + 2
    
    # Auto-adjust columns
    formatting.auto_adjust_column_width(ws, min_width=12, max_width=60)


def create_attendance_teacher_sheet(wb, att_teacher, attendance_df, weekly_stats, 
                                    selected_week_dates, selected_week_id, tier3_df):
    """
    Enhanced attendance teacher sheet with comprehensive analysis
    
    Args:
        wb: Workbook object
        att_teacher: Attendance teacher name
        attendance_df: Full attendance dataframe
        weekly_stats: Weekly aggregated stats
        selected_week_dates: Dates in selected week
        selected_week_id: Week ID
        tier3_df: Tier 3 students dataframe
    """
    from . import attendance_analysis_utils_enhanced as enhanced
    
    sheet_name = att_teacher[:31].replace('/', '-').replace('\\', '-')
    ws = wb.create_sheet(f"AT-{sheet_name}")
    
    # Filter for this attendance teacher
    student_filter = {'AttendanceTeacher': att_teacher}
    att_teacher_data = attendance_df[attendance_df['AttendanceTeacher'] == att_teacher].copy()
    
    current_row = 1
    
    # ===== TITLE =====
    ws.merge_cells(f'A{current_row}:J{current_row}')
    title_cell = ws.cell(row=current_row, column=1)
    title_cell.value = f"Attendance Teacher Report - {att_teacher}"
    title_cell.font = openpyxl.styles.Font(bold=True, size=14)
    current_row += 2
    
    # ===== SUMMARY STATS =====
    att_students = att_teacher_data['StudentID'].unique()
    week_data = att_teacher_data[att_teacher_data['Date'].isin(selected_week_dates)]
    
    total_periods = len(week_data)
    present_periods = len(week_data[week_data['Type'].isin(['present', 'tardy'])])
    attendance_rate = present_periods / total_periods if total_periods > 0 else 0
    
    current_row = formatting.create_summary_box(
        ws, current_row, 1, f"Caseload Summary ({len(att_students)} students)",
        {
            'Attendance Rate This Week': formatting.format_percentage(attendance_rate),
            'Students with Cuts': len(week_data[week_data['cutting?'] == True]['StudentID'].unique()),
            'Students Late to School': len(week_data[week_data['late_to_school?'] == True]['StudentID'].unique()),
            'Total Cut Instances': len(week_data[week_data['cutting?'] == True]),
        }
    )
    current_row += 2
    
    # ===== SECTION 1: TIER 3 STUDENTS =====
    current_row = formatting.create_section_header(
        ws, current_row, "⚠️ TIER 3 STUDENTS - Requiring Immediate Intervention", 10
    )
    
    tier3_caseload = enhanced.get_tier3_students_for_caseload(tier3_df, student_filter)
    
    if len(tier3_caseload) > 0:
        # Format percentages
        tier3_display = tier3_caseload.copy()
        tier3_display['SemesterRate'] = tier3_display['SemesterRate'].apply(formatting.format_percentage)
        tier3_display['RecentRate'] = tier3_display['RecentRate'].apply(formatting.format_percentage)
        tier3_display['Trend'] = tier3_display['Trend'].apply(formatting.format_trend)
        
        for r_idx, row in enumerate(dataframe_to_rows(tier3_display, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(tier3_display.columns))
                else:
                    semester_rate = tier3_caseload.iloc[r_idx - current_row - 1]['SemesterRate']
                    if semester_rate < 0.8:
                        color = 'chronic_absent'
                    elif semester_rate < 0.9:
                        color = 'approaching_chronic'
                    else:
                        color = 'declining_trend'
                    formatting.apply_data_row_style(ws, r_idx, len(tier3_display.columns), color)
        
        current_row += len(tier3_display) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No Tier 3 students in caseload"
        current_row += 2
    
    # ===== SECTION 2: STUDENTS WITH CUTS =====
    current_row = formatting.create_section_header(
        ws, current_row, "🚫 STUDENTS WITH CUTS THIS WEEK", 10
    )
    
    cuts_df = enhanced.get_cuts_by_period_detailed(
        attendance_df, selected_week_dates, student_filter
    )
    
    if len(cuts_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(cuts_df, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(cuts_df.columns))
                else:
                    formatting.apply_data_row_style(ws, r_idx, len(cuts_df.columns), 'cuts')
                    if c_idx == len(cuts_df.columns):  # Last column (PeriodBreakdown)
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        
        current_row += len(cuts_df) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No cuts this week"
        current_row += 2
    
    # ===== SECTION 3: STUDENTS LATE TO SCHOOL =====
    current_row = formatting.create_section_header(
        ws, current_row, "⏰ STUDENTS LATE TO SCHOOL THIS WEEK", 14
    )
    
    late_df = enhanced.get_late_to_school_detailed(
        attendance_df, selected_week_dates, student_filter
    )
    
    if len(late_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(late_df, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(late_df.columns))
                else:
                    trend = late_df.iloc[r_idx - current_row - 1]['Trend']
                    if trend == 'Worsening':
                        color = 'late_to_school'
                    elif trend == 'Improving':
                        color = 'most_improved'
                    else:
                        color = None
                    formatting.apply_data_row_style(ws, r_idx, len(late_df.columns), color)
                    
                    if c_idx == len(late_df.columns):  # Last column
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
        
        current_row += len(late_df) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No students late to school this week"
        current_row += 2
    
    # ===== SECTION 4: STUDENTS ABSENT ALL WEEK =====
    current_row = formatting.create_section_header(
        ws, current_row, "❌ STUDENTS ABSENT ALL WEEK", 10
    )
    
    absent_df = enhanced.get_absent_all_week_with_trends(
        attendance_df, weekly_stats, selected_week_dates, 
        selected_week_id, student_filter
    )
    
    if len(absent_df) > 0:
        absent_display = absent_df.copy()
        absent_display['SemesterRate'] = absent_display['SemesterRate'].apply(formatting.format_percentage)
        absent_display['RecentRate'] = absent_display['RecentRate'].apply(formatting.format_percentage)
        
        for r_idx, row in enumerate(dataframe_to_rows(absent_display, index=False, header=True), current_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == current_row:
                    formatting.apply_header_style(ws, r_idx, len(absent_display.columns))
                else:
                    trend = absent_df.iloc[r_idx - current_row - 1]['Trend']
                    if trend == 'Declining':
                        color = 'chronic_absent'
                    elif trend == 'Improving':
                        color = 'approaching_chronic'
                    else:
                        color = 'chronic_absent'
                    formatting.apply_data_row_style(ws, r_idx, len(absent_display.columns), color)
        
        current_row += len(absent_display) + 2
    else:
        ws.cell(row=current_row, column=1).value = "No students absent all week"
        current_row += 2
    
    # Auto-adjust columns
    formatting.auto_adjust_column_width(ws, min_width=12, max_width=60)

def create_trend_chart(ws, weekly_stats, selected_week_id, start_row):
    """Add a line chart showing attendance trends over the semester"""
    # Prepare data for chart (last 8 weeks)
    chart_data = weekly_stats[
        (weekly_stats['week_id'] <= selected_week_id) &
        (weekly_stats['week_id'] > selected_week_id - 8)
    ].copy()
    
    if len(chart_data) == 0:
        return
    
    # Aggregate by week and year_in_hs
    trend_by_year = chart_data.groupby(['week_id', 'year_in_hs'])['attendance_rate_smooth'].mean().unstack(fill_value=0)
    
    # Write data starting at start_row
    ws.cell(row=start_row, column=1).value = "Week"
    for col_idx, year in enumerate(trend_by_year.columns, 2):
        ws.cell(row=start_row, column=col_idx).value = f"Year {year}"
    
    for row_idx, (week_id, row_data) in enumerate(trend_by_year.iterrows(), start_row + 1):
        ws.cell(row=row_idx, column=1).value = f"Week {week_id}"
        for col_idx, year in enumerate(trend_by_year.columns, 2):
            ws.cell(row=row_idx, column=col_idx).value = row_data[year]
    
    # Create chart
    chart = LineChart()
    chart.title = "Attendance Trend by Year"
    chart.y_axis.title = "Attendance Rate"
    chart.x_axis.title = "Week"
    
    data_ref = Reference(ws, min_col=2, min_row=start_row, 
                        max_col=1 + len(trend_by_year.columns),
                        max_row=start_row + len(trend_by_year))
    cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, 
                        max_row=start_row + len(trend_by_year))
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    ws.add_chart(chart, f"A{start_row + len(trend_by_year) + 2}")