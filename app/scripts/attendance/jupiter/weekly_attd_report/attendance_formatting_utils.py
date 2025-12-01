"""
Excel formatting utilities for attendance analysis
Handles styling, conditional formatting, and visual presentation
"""

import pandas as pd  # ✓ FIXED: Moved to top
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# Color scheme for different attendance issues
COLORS = {
    # Critical issues
    'chronic_absent': 'FFE6E6',      # Light red - >10% absent
    'cuts': 'FFD9D9',                # Red - potential cuts
    'declining_trend': 'FFB3B3',      # Dark red - negative trend
    
    # Warning issues  
    'approaching_chronic': 'FFF4CC', # Yellow - approaching 10%
    'late_to_school': 'FFE6B3',      # Light orange - tardy pattern
    'attendance_error': 'FFD699',    # Orange - data quality issue
    
    # Positive
    'good_attendance': 'E6F4EA',     # Light green - >90% attendance
    'most_improved': 'D4EDDA',       # Green - positive trend
    
    # Neutral/Info
    'missing_data': 'E8E8E8',        # Gray - missing marks
    'header': 'D3D3D3',              # Light gray - headers
    'subheader': 'F0F0F0',           # Very light gray - subheaders
}


def apply_header_style(ws, row_num, num_cols):
    """Apply consistent header styling to a row"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=COLORS['header'], 
                               end_color=COLORS['header'], 
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border


def apply_data_row_style(ws, row_num, num_cols, issue_type=None):
    """
    Apply styling to a data row based on issue type
    
    Args:
        ws: worksheet
        row_num: row number
        num_cols: number of columns
        issue_type: one of 'chronic_absent', 'cuts', 'late_to_school', 
                    'attendance_error', 'declining_trend', etc.
    """
    fill_color = COLORS.get(issue_type, 'FFFFFF')
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = PatternFill(start_color=fill_color, 
                               end_color=fill_color, 
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border


def apply_alternating_rows(ws, start_row, end_row, num_cols, color1='FFFFFF', color2='F8F8F8'):
    """Apply alternating row colors for readability"""
    for row in range(start_row, end_row + 1):
        fill_color = color1 if (row - start_row) % 2 == 0 else color2
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.fill.start_color or cell.fill.start_color.rgb == '00000000':
                cell.fill = PatternFill(start_color=fill_color, 
                                       end_color=fill_color, 
                                       fill_type='solid')


def auto_adjust_column_width(ws, min_width=10, max_width=50):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def add_conditional_formatting_attendance_rate(ws, data_range, rate_column):
    """
    Add conditional formatting for attendance rates
    Green for >90%, Yellow for 80-90%, Red for <80%
    """
    # Good attendance (>90%)
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(
            operator='greaterThan',
            formula=['0.9'],
            fill=PatternFill(start_color=COLORS['good_attendance'], 
                           end_color=COLORS['good_attendance'], 
                           fill_type='solid')
        )
    )
    
    # Approaching chronic (80-90%)
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(
            operator='between',
            formula=['0.8', '0.9'],
            fill=PatternFill(start_color=COLORS['approaching_chronic'], 
                           end_color=COLORS['approaching_chronic'], 
                           fill_type='solid')
        )
    )
    
    # Chronic absent (<80%)
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(
            operator='lessThan',
            formula=['0.8'],
            fill=PatternFill(start_color=COLORS['chronic_absent'], 
                           end_color=COLORS['chronic_absent'], 
                           fill_type='solid')
        )
    )


def format_percentage(value):
    """Format decimal as percentage string"""
    if pd.isna(value):
        return ''
    return f"{value:.1%}"


def format_trend(value):
    """Format trend with arrow indicator"""
    if pd.isna(value):
        return ''
    
    if value > 0.01:
        return f"↑ {value:.3f}"
    elif value < -0.01:
        return f"↓ {value:.3f}"
    else:
        return f"→ {value:.3f}"


def create_section_header(ws, row_num, title, num_cols):
    """Create a section header spanning multiple columns"""
    # Merge cells
    ws.merge_cells(start_row=row_num, start_column=1, 
                   end_row=row_num, end_column=num_cols)
    
    cell = ws.cell(row=row_num, column=1)
    cell.value = title
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    return row_num + 1


def create_summary_box(ws, start_row, start_col, title, stats_dict):
    """
    Create a formatted summary box with statistics
    
    Args:
        ws: worksheet
        start_row: starting row
        start_col: starting column
        title: box title
        stats_dict: dict of {label: value} pairs
    
    Returns:
        end_row: last row used
    """
    current_row = start_row
    
    # Title
    cell = ws.cell(row=current_row, column=start_col)
    cell.value = title
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color=COLORS['subheader'], 
                           end_color=COLORS['subheader'], 
                           fill_type='solid')
    ws.merge_cells(start_row=current_row, start_column=start_col,
                   end_row=current_row, end_column=start_col + 1)
    current_row += 1
    
    # Stats
    for label, value in stats_dict.items():
        label_cell = ws.cell(row=current_row, column=start_col)
        label_cell.value = label
        label_cell.alignment = Alignment(horizontal='right')
        
        value_cell = ws.cell(row=current_row, column=start_col + 1)
        value_cell.value = value
        value_cell.font = Font(bold=True)
        
        current_row += 1
    
    return current_row


def add_legend(ws, start_row, start_col):
    """Add a color legend explaining what each color means"""
    current_row = start_row
    
    # Title
    cell = ws.cell(row=current_row, column=start_col)
    cell.value = "Color Legend"
    cell.font = Font(bold=True, size=11)
    current_row += 1
    
    legend_items = [
        ('Chronic Absence (>10%)', 'chronic_absent'),
        ('Potential Cuts', 'cuts'),
        ('Declining Trend', 'declining_trend'),
        ('Approaching Chronic (8-10%)', 'approaching_chronic'),
        ('Late to School Pattern', 'late_to_school'),
        ('Attendance Error', 'attendance_error'),
        ('Good Attendance (>90%)', 'good_attendance'),
        ('Most Improved', 'most_improved'),
        ('Missing Data', 'missing_data'),
    ]
    
    for label, color_key in legend_items:
        label_cell = ws.cell(row=current_row, column=start_col)
        label_cell.value = label
        label_cell.fill = PatternFill(start_color=COLORS[color_key], 
                                     end_color=COLORS[color_key], 
                                     fill_type='solid')
        current_row += 1
    
    return current_row