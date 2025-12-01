"""
Excel Output Generator Module

Creates comprehensive Excel workbook with teacher impact analysis and student
trajectory analysis, including professional formatting and visualizations.
"""

import pandas as pd
import numpy as np
from io import BytesIO
from typing import Tuple
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule


def generate_excel_output(
    teacher_analysis: Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame],
    student_analysis: Tuple[pd.DataFrame, pd.DataFrame]
) -> BytesIO:
    """
    Generates Excel workbook with teacher and student analysis.
    
    Creates ten sheets:
    DIAGNOSTICS:
    - Teacher Impact Diagnostics (NEW)
    
    TEACHER ANALYSIS:
    - Teacher Impact Summary
    - Teacher Detail
    - Teacher Impact Visualization
    
    STUDENT ANALYSIS (ALL STUDENTS):
    - Student Trajectory Summary
    - Student Term Detail
    - Student Trajectory Visualization
    
    STUDENT ANALYSIS (CURRENT STUDENTS):
    - Current Students Summary
    - Top 10 by Cohort
    - Current Students Visualization
    
    Args:
        teacher_analysis: Tuple of (teacher_summary_df, teacher_detail_df, diagnostics_df)
        student_analysis: Tuple of (student_summary_df, student_term_detail_df)
    
    Returns:
        BytesIO object containing Excel file
    """
    teacher_summary, teacher_detail, teacher_diagnostics = teacher_analysis
    student_summary, student_term_detail = student_analysis
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # DIAGNOSTICS SHEET (FIRST!)
    _create_teacher_diagnostics_sheet(wb, teacher_diagnostics)
    
    # TEACHER ANALYSIS BLOCK
    _create_teacher_summary_sheet(wb, teacher_summary)
    _create_teacher_detail_sheet(wb, teacher_detail)
    _create_teacher_visualization_sheet(wb, teacher_summary)
    
    # STUDENT ANALYSIS BLOCK (ALL STUDENTS)
    _create_student_summary_sheet(wb, student_summary)
    _create_student_term_detail_sheet(wb, student_term_detail)
    _create_student_visualization_sheet(wb, student_summary, student_term_detail)
    
    # CURRENT STUDENTS ANALYSIS (if enrollment data available)
    if 'still_enrolled' in student_summary.columns:
        current_students = student_summary[student_summary['still_enrolled'] == True].copy()
        current_detail = student_term_detail[
            student_term_detail['StudentID'].isin(current_students['StudentID'])
        ].copy()
        
        _create_current_students_summary_sheet(wb, current_students)
        _create_top_10_by_cohort_sheet(wb, current_students)
        _create_current_students_visualization_sheet(wb, current_students)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def _create_teacher_diagnostics_sheet(wb: Workbook, diagnostics_df: pd.DataFrame):
    """Creates diagnostic sheet to help debug teacher impact calculation issues."""
    ws = wb.create_sheet("DIAGNOSTICS - Teacher Impact", 0)  # Make it first sheet
    
    # Title and instructions
    ws['A1'] = "TEACHER IMPACT DIAGNOSTICS"
    ws['A1'].font = Font(size=14, bold=True, color="FF0000")
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Copy the summary statistics below and share with your analyst to debug negative impact issues"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:F2')
    
    # Overall Statistics
    ws['A4'] = "=== OVERALL STATISTICS ==="
    ws['A4'].font = Font(size=12, bold=True)
    
    if not diagnostics_df.empty:
        ws['A5'] = "Total Teacher-Student Pairs Analyzed:"
        ws['B5'] = len(diagnostics_df)
        
        ws['A6'] = "Average Baseline Z-Score:"
        ws['B6'] = round(diagnostics_df['BaselineZScore'].mean(), 4)
        
        ws['A7'] = "Average Post-Teacher Z-Score:"
        ws['B7'] = round(diagnostics_df['PostTeacherZScore'].mean(), 4)
        
        ws['A8'] = "Average Impact:"
        ws['B8'] = round(diagnostics_df['Impact'].mean(), 4)
        
        ws['A9'] = "Median Impact:"
        ws['B9'] = round(diagnostics_df['Impact'].median(), 4)
        
        ws['A10'] = "% Negative Impacts:"
        ws['B10'] = f"{(diagnostics_df['Impact'] < 0).sum() / len(diagnostics_df) * 100:.1f}%"
        
        ws['A11'] = "% Positive Impacts:"
        ws['B11'] = f"{(diagnostics_df['Impact'] > 0).sum() / len(diagnostics_df) * 100:.1f}%"
        
        # Sample size analysis
        ws['A13'] = "=== SAMPLE SIZE ANALYSIS ==="
        ws['A13'].font = Font(size=12, bold=True)
        
        ws['A14'] = "Average Baseline Courses Count:"
        ws['B14'] = round(diagnostics_df['BaselineCoursesCount'].mean(), 2)
        
        ws['A15'] = "Average Post-Teacher Courses Count:"
        ws['B15'] = round(diagnostics_df['PostCoursesCount'].mean(), 2)
        
        ws['A16'] = "Average Terms After Teacher:"
        ws['B16'] = round(diagnostics_df['TermsAfter'].mean(), 2)
        
        # Distribution by baseline courses
        ws['A18'] = "=== IMPACT BY BASELINE SAMPLE SIZE ==="
        ws['A18'].font = Font(size=12, bold=True)
        
        ws['A19'] = "Baseline Courses"
        ws['B19'] = "Count"
        ws['C19'] = "Avg Impact"
        ws['D19'] = "% Negative"
        _format_header_row(ws, start_row=19)
        
        # Group by baseline course count
        for i, baseline_count in enumerate([1, 2, 3, 4, 5, '6-10', '11+'], start=20):
            if baseline_count == '6-10':
                subset = diagnostics_df[(diagnostics_df['BaselineCoursesCount'] >= 6) & 
                                       (diagnostics_df['BaselineCoursesCount'] <= 10)]
            elif baseline_count == '11+':
                subset = diagnostics_df[diagnostics_df['BaselineCoursesCount'] >= 11]
            else:
                subset = diagnostics_df[diagnostics_df['BaselineCoursesCount'] == baseline_count]
            
            if len(subset) > 0:
                ws[f'A{i}'] = str(baseline_count)
                ws[f'B{i}'] = len(subset)
                ws[f'C{i}'] = round(subset['Impact'].mean(), 4)
                ws[f'D{i}'] = f"{(subset['Impact'] < 0).sum() / len(subset) * 100:.1f}%"
        
        # Distribution by teacher term
        ws['A29'] = "=== IMPACT BY TEACHER TERM ==="
        ws['A29'].font = Font(size=12, bold=True)
        
        ws['A30'] = "Teacher Term"
        ws['B30'] = "Count"
        ws['C30'] = "Avg Impact"
        ws['D30'] = "% Negative"
        _format_header_row(ws, start_row=30)
        
        term_groups = diagnostics_df.groupby('TeacherTerm')['Impact'].agg(['count', 'mean', lambda x: (x < 0).sum() / len(x) * 100]).reset_index()
        term_groups.columns = ['TeacherTerm', 'Count', 'AvgImpact', 'PctNegative']
        term_groups = term_groups.sort_values('TeacherTerm')
        
        for i, (_, row) in enumerate(term_groups.head(20).iterrows(), start=31):  # Show first 20 terms
            ws[f'A{i}'] = row['TeacherTerm']
            ws[f'B{i}'] = row['Count']
            ws[f'C{i}'] = round(row['AvgImpact'], 4)
            ws[f'D{i}'] = f"{row['PctNegative']:.1f}%"
        
        # Sample records for detailed inspection
        ws['F4'] = "=== SAMPLE RECORDS (First 20) ==="
        ws['F4'].font = Font(size=12, bold=True)
        
        sample_cols = ['Teacher', 'StudentID', 'TeacherTerm', 'BaselineZScore', 
                       'PostTeacherZScore', 'Impact', 'BaselineCoursesCount', 'PostCoursesCount']
        
        for col_idx, col in enumerate(sample_cols, start=6):  # Start at column F
            ws.cell(row=5, column=col_idx, value=col)
        _format_header_row(ws, start_row=5, start_col=6)
        
        for row_idx, (_, record) in enumerate(diagnostics_df.head(20).iterrows(), start=6):
            for col_idx, col in enumerate(sample_cols, start=6):
                value = record[col]
                if isinstance(value, (int, float)) and col in ['BaselineZScore', 'PostTeacherZScore', 'Impact']:
                    value = round(value, 4)
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    else:
        ws['A5'] = "No diagnostic data available"
    
    # Auto-size columns
    _auto_size_columns(ws)


def _create_teacher_summary_sheet(wb: Workbook, teacher_summary: pd.DataFrame):
    """Creates Teacher Impact Summary sheet with peer comparison data."""
    ws = wb.create_sheet("Teacher Impact Summary")
    
    # Prepare display columns
    display_cols = [
        'Teacher', 'TotalStudents', 'AvgBaselineZScore', 'AvgPostTeacherZScore',
        'AvgPeerPostZScore', 'OverallImpact', 'SameContentImpact', 'CrossContentImpact',
        'Confidence'
    ]
    
    display_df = teacher_summary[display_cols].copy() if not teacher_summary.empty else pd.DataFrame(columns=display_cols)
    
    # Round numeric columns (z-scores need precision)
    numeric_cols = ['AvgBaselineZScore', 'AvgPostTeacherZScore', 'AvgPeerPostZScore', 
                    'OverallImpact', 'SameContentImpact', 'CrossContentImpact']
    for col in numeric_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].round(3)
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), 1):
        ws.append(row)
    
    # Format header
    _format_header_row(ws)
    
    # Apply conditional formatting to impact columns (z-score scale)
    if len(display_df) > 0:
        from openpyxl.utils import get_column_letter
        impact_cols = ['OverallImpact', 'SameContentImpact', 'CrossContentImpact']
        for col in impact_cols:
            if col in display_cols:
                col_letter = get_column_letter(display_cols.index(col) + 1)
                _apply_z_score_conditional_formatting(ws, col_letter, 2, len(display_df) + 1)
    
    # Auto-size columns
    _auto_size_columns(ws)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Enable filters
    ws.auto_filter.ref = ws.dimensions


def _create_teacher_detail_sheet(wb: Workbook, teacher_detail: pd.DataFrame):
    """Creates Teacher Detail sheet with student-level peer comparison data."""
    ws = wb.create_sheet("Teacher Detail")
    
    # Prepare display columns
    base_cols = [
        'Teacher', 'StudentID', 'TeacherTerm', 'ContentArea',
        'BaselineZScore', 'PostTeacherZScore', 'PeerPostZScore', 'OverallImpact',
        'SameContentImpact', 'CrossContentImpact', 'TermsAnalyzed'
    ]
    
    # Add student info columns if available
    optional_cols = ['LastName', 'FirstName', 'GEC', 'still_enrolled']
    display_cols = base_cols.copy()
    for col in optional_cols:
        if col in teacher_detail.columns:
            # Insert student info after StudentID
            display_cols.insert(2, col)
    
    display_df = teacher_detail[display_cols].copy() if not teacher_detail.empty else pd.DataFrame(columns=display_cols)
    
    # Round numeric columns (z-scores need precision)
    numeric_cols = ['BaselineZScore', 'PostTeacherZScore', 'PeerPostZScore', 'OverallImpact',
                    'SameContentImpact', 'CrossContentImpact']
    for col in numeric_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].round(3)
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), 1):
        ws.append(row)
    
    # Format header
    _format_header_row(ws)
    
    # Apply conditional formatting to impact columns
    if len(display_df) > 0:
        from openpyxl.utils import get_column_letter
        impact_cols = ['OverallImpact', 'SameContentImpact', 'CrossContentImpact']
        for col in impact_cols:
            if col in display_cols:
                col_letter = get_column_letter(display_cols.index(col) + 1)
                _apply_z_score_conditional_formatting(ws, col_letter, 2, len(display_df) + 1)
    
    # Auto-size columns
    _auto_size_columns(ws)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Enable filters
    ws.auto_filter.ref = ws.dimensions


def _create_teacher_visualization_sheet(wb: Workbook, teacher_summary: pd.DataFrame):
    """Creates Teacher Impact Visualization sheet with charts."""
    ws = wb.create_sheet("Teacher Impact Viz")
    
    if teacher_summary.empty or len(teacher_summary) == 0:
        ws['A1'] = "Insufficient data for visualization"
        return
    
    # Add title
    ws['A1'] = "Teacher Impact Analysis - Z-Score Based Performance"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Top 10 teachers by overall impact
    ws['A3'] = "Top 10 Teachers by Overall Z-Score Impact"
    ws['A3'].font = Font(size=12, bold=True)
    
    top_10 = teacher_summary.nlargest(10, 'OverallImpact')[['Teacher', 'OverallImpact']].copy()
    
    # Write top 10 data
    ws['A4'] = "Teacher"
    ws['B4'] = "Overall Impact (Z-Score)"
    _format_header_row(ws, start_row=4)
    
    for idx, (_, row) in enumerate(top_10.iterrows(), start=5):
        ws[f'A{idx}'] = row['Teacher']
        ws[f'B{idx}'] = round(row['OverallImpact'], 3) if pd.notna(row['OverallImpact']) else 0
    
    # Create bar chart
    chart = BarChart()
    chart.title = "Top 10 Teachers by Overall Z-Score Impact"
    chart.x_axis.title = "Teacher"
    chart.y_axis.title = "Impact (Z-Score Change)"
    
    data = Reference(ws, min_col=2, min_row=4, max_row=4 + len(top_10))
    cats = Reference(ws, min_col=1, min_row=5, max_row=4 + len(top_10))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, "D4")
    
    # Impact distribution
    ws['A20'] = "Z-Score Impact Distribution"
    ws['A20'].font = Font(size=12, bold=True)
    
    ws['A21'] = "Metric"
    ws['B21'] = "Mean"
    ws['C21'] = "Median"
    ws['D21'] = "Std Dev"
    _format_header_row(ws, start_row=21)
    
    ws['A22'] = "Overall Impact"
    ws['B22'] = round(teacher_summary['OverallImpact'].mean(), 3)
    ws['C22'] = round(teacher_summary['OverallImpact'].median(), 3)
    ws['D22'] = round(teacher_summary['OverallImpact'].std(), 3)
    
    ws['A23'] = "Same Content Impact"
    ws['B23'] = round(teacher_summary['SameContentImpact'].mean(), 3) if teacher_summary['SameContentImpact'].notna().any() else 0
    ws['C23'] = round(teacher_summary['SameContentImpact'].median(), 3) if teacher_summary['SameContentImpact'].notna().any() else 0
    ws['D23'] = round(teacher_summary['SameContentImpact'].std(), 3) if teacher_summary['SameContentImpact'].notna().any() else 0
    
    ws['A24'] = "Cross Content Impact"
    ws['B24'] = round(teacher_summary['CrossContentImpact'].mean(), 3) if teacher_summary['CrossContentImpact'].notna().any() else 0
    ws['C24'] = round(teacher_summary['CrossContentImpact'].median(), 3) if teacher_summary['CrossContentImpact'].notna().any() else 0
    ws['D24'] = round(teacher_summary['CrossContentImpact'].std(), 3) if teacher_summary['CrossContentImpact'].notna().any() else 0
    
    # Interpretation guide
    ws['A27'] = "Z-Score Impact Interpretation Guide"
    ws['A27'].font = Font(size=11, bold=True)
    
    ws['A28'] = "Impact Range"
    ws['B28'] = "Interpretation"
    _format_header_row(ws, start_row=28)
    
    ws['A29'] = "+0.5 or higher"
    ws['B29'] = "Strong positive impact - students improve significantly"
    
    ws['A30'] = "+0.2 to +0.5"
    ws['B30'] = "Moderate positive impact"
    
    ws['A31'] = "-0.2 to +0.2"
    ws['B31'] = "Neutral impact (within normal variation)"
    
    ws['A32'] = "-0.5 to -0.2"
    ws['B32'] = "Moderate negative impact"
    
    ws['A33'] = "-0.5 or lower"
    ws['B33'] = "Strong negative impact - students decline after course"


def _create_student_summary_sheet(wb: Workbook, student_summary: pd.DataFrame):
    """Creates Student Trajectory Summary sheet with z-score metrics."""
    ws = wb.create_sheet("Student Trajectory Summary")
    
    # Prepare display columns
    base_cols = [
        'StudentID', 'FirstTerm', 'LastTerm', 'TermsAnalyzed',
        'TrajectorySlope', 'AvgZScore', 'MostRecentZScore',
        'Trend', 'TermsAboveAvg', 'TermsBelowAvg'
    ]
    
    # Add student info columns if available
    optional_cols = ['LastName', 'FirstName', 'GEC', 'still_enrolled']
    display_cols = base_cols.copy()
    for col in optional_cols:
        if col in student_summary.columns:
            # Insert student info after StudentID
            display_cols.insert(1, col)
    
    display_df = student_summary[display_cols].copy() if not student_summary.empty else pd.DataFrame(columns=display_cols)
    
    # Round numeric columns
    numeric_cols = ['TrajectorySlope', 'AvgZScore', 'MostRecentZScore']
    for col in numeric_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].round(3)  # Z-scores need more precision
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), 1):
        ws.append(row)
    
    # Format header
    _format_header_row(ws)
    
    # Apply conditional formatting
    if len(display_df) > 0:
        from openpyxl.utils import get_column_letter
        if 'TrajectorySlope' in display_cols:
            slope_col = get_column_letter(display_cols.index('TrajectorySlope') + 1)
            _apply_impact_conditional_formatting(ws, slope_col, 2, len(display_df) + 1)
        if 'AvgZScore' in display_cols:
            avg_col = get_column_letter(display_cols.index('AvgZScore') + 1)
            _apply_impact_conditional_formatting(ws, avg_col, 2, len(display_df) + 1)
        if 'MostRecentZScore' in display_cols:
            recent_col = get_column_letter(display_cols.index('MostRecentZScore') + 1)
            _apply_impact_conditional_formatting(ws, recent_col, 2, len(display_df) + 1)
    
    # Auto-size columns
    _auto_size_columns(ws)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Enable filters
    ws.auto_filter.ref = ws.dimensions


def _create_student_term_detail_sheet(wb: Workbook, student_term_detail: pd.DataFrame):
    """Creates Student Term Detail sheet with z-score data."""
    ws = wb.create_sheet("Student Term Detail")
    
    # Prepare display columns
    base_cols = [
        'StudentID', 'Year', 'TermNum', 'Term', 'AvgZScore'
    ]
    
    # Add student info columns if available
    optional_cols = ['LastName', 'FirstName', 'GEC', 'still_enrolled']
    display_cols = base_cols.copy()
    for col in optional_cols:
        if col in student_term_detail.columns:
            # Insert student info after StudentID
            display_cols.insert(1, col)
    
    display_df = student_term_detail[display_cols].copy() if not student_term_detail.empty else pd.DataFrame(columns=display_cols)
    
    # Round numeric columns
    if 'AvgZScore' in display_df.columns:
        display_df['AvgZScore'] = display_df['AvgZScore'].round(3)  # Z-scores need precision
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), 1):
        ws.append(row)
    
    # Format header
    _format_header_row(ws)
    
    # Apply conditional formatting to z-score
    if len(display_df) > 0:
        from openpyxl.utils import get_column_letter
        if 'AvgZScore' in display_cols:
            z_col = get_column_letter(display_cols.index('AvgZScore') + 1)
            _apply_z_score_conditional_formatting(ws, z_col, 2, len(display_df) + 1)
    
    # Auto-size columns
    _auto_size_columns(ws)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Enable filters
    ws.auto_filter.ref = ws.dimensions


def _create_student_visualization_sheet(wb: Workbook, student_summary: pd.DataFrame, 
                                       student_term_detail: pd.DataFrame):
    """Creates Student Trajectory Visualization sheet with z-score analysis."""
    ws = wb.create_sheet("Student Trajectory Viz")
    
    if student_summary.empty or len(student_summary) == 0:
        ws['A1'] = "Insufficient data for visualization"
        return
    
    # Add title
    ws['A1'] = "Student Trajectory Analysis - Z-Score Performance"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Top improvers by trajectory slope
    ws['A3'] = "Top 10 Improving Students (by Trajectory Slope)"
    ws['A3'].font = Font(size=12, bold=True)
    
    top_10 = student_summary.nlargest(10, 'TrajectorySlope')[['StudentID', 'TrajectorySlope', 'AvgZScore']].copy()
    
    ws['A4'] = "Student ID"
    ws['B4'] = "Trajectory Slope"
    ws['C4'] = "Avg Z-Score"
    _format_header_row(ws, start_row=4)
    
    for idx, (_, row) in enumerate(top_10.iterrows(), start=5):
        ws[f'A{idx}'] = row['StudentID']
        ws[f'B{idx}'] = round(row['TrajectorySlope'], 3) if pd.notna(row['TrajectorySlope']) else 0
        ws[f'C{idx}'] = round(row['AvgZScore'], 3) if pd.notna(row['AvgZScore']) else 0
    
    # Bottom performers (declining)
    ws['A17'] = "Top 10 Declining Students (by Trajectory Slope)"
    ws['A17'].font = Font(size=12, bold=True)
    
    bottom_10 = student_summary.nsmallest(10, 'TrajectorySlope')[['StudentID', 'TrajectorySlope', 'AvgZScore']].copy()
    
    ws['A18'] = "Student ID"
    ws['B18'] = "Trajectory Slope"
    ws['C18'] = "Avg Z-Score"
    _format_header_row(ws, start_row=18)
    
    for idx, (_, row) in enumerate(bottom_10.iterrows(), start=19):
        ws[f'A{idx}'] = row['StudentID']
        ws[f'B{idx}'] = round(row['TrajectorySlope'], 3) if pd.notna(row['TrajectorySlope']) else 0
        ws[f'C{idx}'] = round(row['AvgZScore'], 3) if pd.notna(row['AvgZScore']) else 0
    
    # Trend distribution
    ws['E3'] = "Trend Distribution"
    ws['E3'].font = Font(size=12, bold=True)
    
    ws['E4'] = "Trend"
    ws['F4'] = "Count"
    ws['G4'] = "Percentage"
    _format_header_row(ws, start_row=4, start_col=5)
    
    trend_counts = student_summary['Trend'].value_counts()
    total = len(student_summary)
    
    for idx, (trend, count) in enumerate(trend_counts.items(), start=5):
        ws[f'E{idx}'] = trend
        ws[f'F{idx}'] = count
        ws[f'G{idx}'] = f"{(count/total*100):.1f}%"
    
    # Z-Score distribution statistics
    ws['E15'] = "Z-Score Performance Statistics"
    ws['E15'].font = Font(size=12, bold=True)
    
    ws['E16'] = "Metric"
    ws['F16'] = "Value"
    _format_header_row(ws, start_row=16, start_col=5)
    
    ws['E17'] = "Avg Trajectory Slope"
    ws['F17'] = round(student_summary['TrajectorySlope'].mean(), 3)
    
    ws['E18'] = "Median Trajectory Slope"
    ws['F18'] = round(student_summary['TrajectorySlope'].median(), 3)
    
    ws['E19'] = "Avg Z-Score (All Students)"
    ws['F19'] = round(student_summary['AvgZScore'].mean(), 3)
    
    ws['E20'] = "Students Above Avg (Z > 0)"
    ws['F20'] = len(student_summary[student_summary['AvgZScore'] > 0])
    
    ws['E21'] = "Students Below Avg (Z < 0)"
    ws['F21'] = len(student_summary[student_summary['AvgZScore'] < 0])


def _format_header_row(ws, start_row=1, start_col=1):
    """Applies formatting to header row."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[start_row]:
        if cell.column >= start_col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')


def _apply_impact_conditional_formatting(ws, column, start_row, end_row):
    """Applies color scale conditional formatting to impact columns."""
    # Color scale: red (negative) -> white (zero) -> green (positive)
    color_scale = ColorScaleRule(
        start_type='num', start_value=-10, start_color='F8696B',
        mid_type='num', mid_value=0, mid_color='FFFFFF',
        end_type='num', end_value=10, end_color='63BE7B'
    )
    
    ws.conditional_formatting.add(
        f'{column}{start_row}:{column}{end_row}',
        color_scale
    )


def _apply_z_score_conditional_formatting(ws, column, start_row, end_row):
    """Applies color scale conditional formatting specifically for z-scores."""
    # Z-score color scale: red (-3) -> white (0) -> green (+3)
    # Z-scores typically range from -3 to +3 (3 standard deviations)
    color_scale = ColorScaleRule(
        start_type='num', start_value=-3, start_color='F8696B',
        mid_type='num', mid_value=0, mid_color='FFFFFF',
        end_type='num', end_value=3, end_color='63BE7B'
    )
    
    ws.conditional_formatting.add(
        f'{column}{start_row}:{column}{end_row}',
        color_scale
    )


def _auto_size_columns(ws, min_width=10, max_width=50):
    """Auto-sizes columns based on content."""
    from openpyxl.cell.cell import MergedCell
    
    for column in ws.columns:
        max_length = 0
        column_letter = None
        
        for cell in column:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            
            # Get column letter from first non-merged cell
            if column_letter is None:
                column_letter = cell.column_letter
            
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        # Only adjust if we found a valid column letter
        if column_letter:
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width


def _create_current_students_summary_sheet(wb: Workbook, current_students: pd.DataFrame):
    """Creates summary sheet for currently enrolled students only."""
    ws = wb.create_sheet("Current Students Summary")
    
    # Prepare display columns
    display_cols = [
        'StudentID', 'LastName', 'FirstName', 'GEC', 'FirstTerm', 'LastTerm',
        'TermsAnalyzed', 'TrajectorySlope', 'AvgZScore', 'MostRecentZScore',
        'Trend', 'TermsAboveAvg', 'TermsBelowAvg'
    ]
    
    # Filter to only include columns that exist
    display_cols = [col for col in display_cols if col in current_students.columns]
    display_df = current_students[display_cols].copy() if not current_students.empty else pd.DataFrame(columns=display_cols)
    
    # Round numeric columns
    numeric_cols = ['TrajectorySlope', 'AvgZScore', 'MostRecentZScore']
    for col in numeric_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].round(3)
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), 1):
        ws.append(row)
    
    # Format header
    _format_header_row(ws)
    
    # Apply conditional formatting
    if len(display_df) > 0:
        from openpyxl.utils import get_column_letter
        if 'TrajectorySlope' in display_cols:
            slope_col = get_column_letter(display_cols.index('TrajectorySlope') + 1)
            _apply_impact_conditional_formatting(ws, slope_col, 2, len(display_df) + 1)
        if 'AvgZScore' in display_cols:
            avg_col = get_column_letter(display_cols.index('AvgZScore') + 1)
            _apply_z_score_conditional_formatting(ws, avg_col, 2, len(display_df) + 1)
        if 'MostRecentZScore' in display_cols:
            recent_col = get_column_letter(display_cols.index('MostRecentZScore') + 1)
            _apply_z_score_conditional_formatting(ws, recent_col, 2, len(display_df) + 1)
    
    # Auto-size columns
    _auto_size_columns(ws)
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Enable filters
    ws.auto_filter.ref = ws.dimensions


def _create_top_10_by_cohort_sheet(wb: Workbook, current_students: pd.DataFrame):
    """Creates sheet showing top 10 performers by cohort (GEC)."""
    ws = wb.create_sheet("Top 10 by Cohort")
    
    if current_students.empty or 'GEC' not in current_students.columns:
        ws['A1'] = "Insufficient data for cohort analysis"
        return
    
    # Add title
    ws['A1'] = "Top 10 Improving Students by Cohort (by Trajectory Slope)"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:I1')
    
    # Get unique cohorts and sort
    cohorts = sorted(current_students['GEC'].dropna().unique())
    
    current_row = 3
    
    for cohort in cohorts:
        # Filter to this cohort
        cohort_students = current_students[current_students['GEC'] == cohort].copy()
        
        # Get top 10 by TrajectorySlope (rate of improvement)
        top_10 = cohort_students.nlargest(10, 'TrajectorySlope')
        
        if len(top_10) == 0:
            continue
        
        # Add cohort header
        ws[f'A{current_row}'] = f"Cohort {cohort}"
        ws[f'A{current_row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{current_row}:I{current_row}')
        current_row += 1
        
        # Add column headers
        headers = ['Rank', 'StudentID', 'LastName', 'FirstName', 'TrajectorySlope', 
                   'AvgZScore', 'MostRecentZScore', 'TermsAnalyzed', 'Trend']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        current_row += 1
        
        # Add top 10 data
        for rank, (_, student) in enumerate(top_10.iterrows(), start=1):
            ws.cell(row=current_row, column=1, value=rank)
            ws.cell(row=current_row, column=2, value=student['StudentID'])
            ws.cell(row=current_row, column=3, value=student.get('LastName', ''))
            ws.cell(row=current_row, column=4, value=student.get('FirstName', ''))
            ws.cell(row=current_row, column=5, value=round(student['TrajectorySlope'], 3) if pd.notna(student['TrajectorySlope']) else '')
            ws.cell(row=current_row, column=6, value=round(student['AvgZScore'], 3) if pd.notna(student['AvgZScore']) else '')
            ws.cell(row=current_row, column=7, value=round(student['MostRecentZScore'], 3) if pd.notna(student['MostRecentZScore']) else '')
            ws.cell(row=current_row, column=8, value=student.get('TermsAnalyzed', ''))
            ws.cell(row=current_row, column=9, value=student.get('Trend', ''))
            current_row += 1
        
        # Add spacing between cohorts
        current_row += 2
    
    # Auto-size columns
    _auto_size_columns(ws)


def _create_current_students_visualization_sheet(wb: Workbook, current_students: pd.DataFrame):
    """Creates visualization sheet for currently enrolled students."""
    ws = wb.create_sheet("Current Students Viz")
    
    if current_students.empty:
        ws['A1'] = "Insufficient data for visualization"
        return
    
    # Add title
    ws['A1'] = "Current Students Analysis - Z-Score Performance"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Cohort performance comparison
    if 'GEC' in current_students.columns:
        ws['A3'] = "Average Trajectory Slope by Cohort"
        ws['A3'].font = Font(size=12, bold=True)
        
        cohort_avg = current_students.groupby('GEC')['TrajectorySlope'].mean().reset_index()
        cohort_avg = cohort_avg.sort_values('GEC')
        
        ws['A4'] = "Cohort"
        ws['B4'] = "Avg Trajectory Slope"
        ws['C4'] = "Avg Z-Score"
        ws['D4'] = "Student Count"
        _format_header_row(ws, start_row=4)
        
        cohort_counts = current_students.groupby('GEC').size().to_dict()
        cohort_z_scores = current_students.groupby('GEC')['AvgZScore'].mean().to_dict()
        
        for idx, (_, row) in enumerate(cohort_avg.iterrows(), start=5):
            cohort = row['GEC']
            ws[f'A{idx}'] = cohort
            ws[f'B{idx}'] = round(row['TrajectorySlope'], 3) if pd.notna(row['TrajectorySlope']) else 0
            ws[f'C{idx}'] = round(cohort_z_scores.get(cohort, 0), 3)
            ws[f'D{idx}'] = cohort_counts.get(cohort, 0)
    
    # Trend distribution
    ws['F3'] = "Trend Distribution (Current Students)"
    ws['F3'].font = Font(size=12, bold=True)
    
    ws['F4'] = "Trend"
    ws['G4'] = "Count"
    ws['H4'] = "Percentage"
    _format_header_row(ws, start_row=4, start_col=6)
    
    trend_counts = current_students['Trend'].value_counts()
    total_students = len(current_students)
    
    for idx, (trend, count) in enumerate(trend_counts.items(), start=5):
        ws[f'F{idx}'] = trend
        ws[f'G{idx}'] = count
        ws[f'H{idx}'] = f"{(count/total_students*100):.1f}%"
    
    # Top performers
    row_offset = len(trend_counts) + 7
    ws[f'F{row_offset}'] = "Top 10 Current Students (by Trajectory Slope)"
    ws[f'F{row_offset}'].font = Font(size=12, bold=True)
    
    top_10_overall = current_students.nlargest(10, 'TrajectorySlope')[
        ['StudentID', 'LastName', 'FirstName', 'GEC', 'TrajectorySlope', 'AvgZScore']
    ].copy()
    
    ws[f'F{row_offset + 1}'] = "Student ID"
    ws[f'G{row_offset + 1}'] = "Last Name"
    ws[f'H{row_offset + 1}'] = "First Name"
    ws[f'I{row_offset + 1}'] = "Cohort"
    ws[f'J{row_offset + 1}'] = "Trajectory Slope"
    ws[f'K{row_offset + 1}'] = "Avg Z-Score"
    _format_header_row(ws, start_row=row_offset + 1, start_col=6)
    
    for idx, (_, student) in enumerate(top_10_overall.iterrows(), start=row_offset + 2):
        ws[f'F{idx}'] = student['StudentID']
        ws[f'G{idx}'] = student.get('LastName', '')
        ws[f'H{idx}'] = student.get('FirstName', '')
        ws[f'I{idx}'] = student.get('GEC', '')
        ws[f'J{idx}'] = round(student['TrajectorySlope'], 3) if pd.notna(student['TrajectorySlope']) else 0
        ws[f'K{idx}'] = round(student['AvgZScore'], 3) if pd.notna(student['AvgZScore']) else 0